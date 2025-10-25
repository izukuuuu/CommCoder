# -*- coding: utf-8 -*-
"""
分类人工调整 Web GUI（持久化 Session + 服务端保存 Excel + 内存优化）
运行：
  pip install fastapi uvicorn[standard] pandas openpyxl requests python-multipart scikit-learn sentence-transformers python-dotenv
  python app.py
  打开 http://127.0.0.1:8000
"""
import base64
import io, os, gc, re, json, uuid, time, math, traceback, hashlib, random
from collections import Counter, defaultdict
from datetime import datetime
from threading import Lock
from typing import Dict, Any, List, Optional, Tuple, Set

import numpy as np
import pandas as pd
import requests
from fastapi import FastAPI, UploadFile, File, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse, HTMLResponse
from sklearn.cluster import KMeans
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.decomposition import PCA
from sklearn.feature_extraction.text import TfidfVectorizer
from wordcloud import WordCloud

try:
    import nltk  # type: ignore
    from nltk.corpus import stopwords as nltk_stopwords  # type: ignore
except Exception:  # pragma: no cover - 可选依赖
    nltk = None  # type: ignore
    nltk_stopwords = None  # type: ignore

try:
    import hdbscan  # type: ignore
except Exception:  # pragma: no cover - 可选依赖
    hdbscan = None
from dotenv import load_dotenv, set_key
import uvicorn
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties

# ============== 可选：sentence-transformers ==============
_ST_MODEL = None
_ST_MODEL_NAME = None
_ST_MODEL_FALLBACKS = [
    "text2vec-base-multilingual",
    "paraphrase-multilingual-MiniLM-L12-v2",
    "uer/sbert-base-chinese-nli",
    "all-MiniLM-L6-v2",
]
try:
    from sentence_transformers import SentenceTransformer
except Exception:
    SentenceTransformer = None

load_dotenv(dotenv_path=".env")

# ================== 常量与目录 ==================
plt.rcParams["font.family"] = ["Times New Roman", "Times", "serif"]
plt.rcParams["axes.labelcolor"] = "#1f2937"
plt.rcParams["xtick.color"] = "#1f2937"
plt.rcParams["ytick.color"] = "#1f2937"
plt.rcParams["axes.unicode_minus"] = False

APP_VERSION = "4.0.0"
TIMELINE_MIN_YEAR = 1965
SESS_DIR    = os.getenv("SESS_DIR", "sessions")
OUTPUT_DIR  = os.getenv("OUTPUT_DIR", "Output")
STATIC_DIR  = "static"
SAMPLING_DIR_NAME = "sampling_audits"
SAMPLING_DEFAULT_RATE = float(os.getenv("SAMPLING_DEFAULT_RATE", "0.03"))

# 限制常驻内存中的 DataFrame 会话数量，LRU 超限自动卸载（磁盘已持久化，不丢）
MAX_MEMORY_SESSIONS = int(os.getenv("MAX_MEMORY_SESSIONS", "3"))

os.makedirs(SESS_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(STATIC_DIR, exist_ok=True)

# ============== 全局内存（LRU + 锁） ==============
DATASTORE: Dict[str, pd.DataFrame] = {}
PROGRESS: Dict[str, Dict[str, Any]] = {}
PROGRESS_LOCK = Lock()

SESSION_LOCKS: Dict[str, Lock] = {}
SESSION_TOUCH: Dict[str, float] = {}  # 最近访问时间戳（LRU）
SESSION_LOCKS_LOCK = Lock()

SCATTER_CACHE: Dict[str, Dict[str, Any]] = {}
TOPIC_VIZ_CACHE: Dict[str, Dict[str, Any]] = {}

SUMMARY_COLUMN = "结构化总结"
KEYWORD_TEXT_COLUMNS = ["Article Title", "Abstract"]
TEXT_COLUMNS = [SUMMARY_COLUMN] + KEYWORD_TEXT_COLUMNS
TOKEN_PATTERN = re.compile(r"[A-Za-z]{2,}|[\u4e00-\u9fa5]{2,}|\d{2,}")
YEAR_PATTERN = re.compile(r"(18|19|20|21)\d{2}")

ADJUST_TOPIC_COL = "研究主题（议题）分类_调整"
ADJUST_FIELD_COL = "研究领域分类_调整"
RANK_SCORE_COL = "智能排序分数"
RANK_GROUP_COL = "智能排序分组"
RANK_OUTLIER_COL = "智能排序_离群"
RANK_ALGO_COL = "智能排序算法"


def _load_english_stopwords() -> Set[str]:
    if nltk_stopwords is None:
        return set()
    try:
        return set(nltk_stopwords.words("english"))
    except LookupError:
        if nltk is not None:
            try:
                nltk.download("stopwords", quiet=True)
                return set(nltk_stopwords.words("english"))
            except Exception:
                return set()
        return set()
    except Exception:
        return set()


ENGLISH_STOPWORDS = _load_english_stopwords()
BASE_STOPWORDS = set(
    [
        "the",
        "and",
        "with",
        "from",
        "that",
        "this",
        "which",
        "were",
        "have",
        "has",
        "research",
        "study",
        "analysis",
        "based",
        "results",
        "data",
        "using",
        "into",
        "对于",
        "以及",
        "研究",
        "方法",
        "结果",
        "分析",
        "提出",
        "进行",
        "领域",
        "主题",
        "意义",
        "探讨",
        "关注",
        "问题",
        "关系",
        "影响",
        "发展",
        "中国",
        "美国",
        "社会",
        "文章",
        "作者",
        "指出",
    ]
)
STOPWORDS = BASE_STOPWORDS.union(ENGLISH_STOPWORDS)

WORDCLOUD_FONT_ENV = os.getenv("WORDCLOUD_FONT_PATH", "").strip()
WORDCLOUD_FONT_CANDIDATES = [
    WORDCLOUD_FONT_ENV or "",
    "C:/Windows/Fonts/simsunb.ttf",
    "C:/Windows/Fonts/simsun.ttc",
    "C:/Windows/Fonts/simfang.ttf",
    "C:/Windows/Fonts/simkai.ttf",
    "C:/Windows/Fonts/STZHONGS.TTF",
    "C:/Windows/Fonts/timesbd.ttf",
    "C:/Windows/Fonts/msyhbd.ttc",
    "C:/Windows/Fonts/msyh.ttc",
    "C:/Windows/Fonts/simhei.ttf",
    "/System/Library/Fonts/STHeiti Light.ttc",
    "/System/Library/Fonts/Songti.ttc",
    "/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc",
    "/usr/share/fonts/truetype/noto/NotoSerifCJK-Regular.ttc",
    "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
]


def _resolve_wordcloud_font() -> Optional[str]:
    """尝试选择一款偏粗的衬线字体，优先使用 Windows 自带字体。"""

    for path in WORDCLOUD_FONT_CANDIDATES:
        if not path:
            continue
        if os.path.exists(path):
            return path
    return None


WORDCLOUD_FONT_PATH = _resolve_wordcloud_font()
WORDCLOUD_WIDTH = int(os.getenv("WORDCLOUD_WIDTH", "1600"))
WORDCLOUD_HEIGHT = int(os.getenv("WORDCLOUD_HEIGHT", "900"))

SCATTER_COLORS = [
    "#2563eb",
    "#1d4ed8",
    "#0ea5e9",
    "#10b981",
    "#f97316",
    "#f43f5e",
    "#a855f7",
    "#facc15",
    "#14b8a6",
    "#6366f1",
]

SCATTER_CHINESE_FONT: Optional[FontProperties] = None


def _resolve_scatter_chinese_font() -> Optional[FontProperties]:
    """优先复用词云字体，失败时按常见家族名称回退。"""

    candidates: List[str] = []
    if WORDCLOUD_FONT_PATH:
        candidates.append(WORDCLOUD_FONT_PATH)
    candidates.extend(
        [
            "C:/Windows/Fonts/simsun.ttc",
            "C:/Windows/Fonts/simfang.ttf",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/msyh.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/noto/NotoSerifCJK-Regular.ttc",
        ]
    )
    for path in candidates:
        if not path:
            continue
        if not os.path.exists(path):
            continue
        try:
            return FontProperties(fname=path)
        except Exception:
            continue
    for family in ["SimSun", "SimHei", "Songti SC", "Microsoft YaHei", "Noto Sans CJK SC"]:
        try:
            font_path = matplotlib.font_manager.findfont(FontProperties(family=family), fallback_to_default=False)
            if font_path:
                return FontProperties(fname=font_path)
        except Exception:
            continue
    return None


SCATTER_CHINESE_FONT = _resolve_scatter_chinese_font()


def _contains_cjk(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text or ""))


def _scatter_color_for_cluster(cluster: int, has_clusters: bool) -> str:
    if not has_clusters:
        return "#2563eb"
    if cluster < 0:
        return "#ef4444"
    idx = abs(int(cluster)) % len(SCATTER_COLORS)
    return SCATTER_COLORS[idx]


def _format_pipeline(vectorizer: str, reduction: str) -> str:
    parts = []
    if vectorizer:
        parts.append(vectorizer.upper())
    if reduction:
        parts.append(reduction.upper())
    return " → ".join(parts)


def _scatter_summary(
    cluster_count: int,
    cluster_algo: str,
    explained: float,
    pipeline_text: str,
    total_docs: int,
) -> str:
    summary_parts: List[str] = []
    if cluster_count <= 1:
        summary_parts.append("未进一步聚类")
    else:
        algo = cluster_algo.upper() if cluster_algo else ""
        if algo:
            summary_parts.append(f"{algo} 聚 {cluster_count} 簇")
        else:
            summary_parts.append(f"聚 {cluster_count} 簇")
    summary_parts.append(f"方差解释 {explained * 100:.1f}%")
    if pipeline_text:
        summary_parts.append(f"语义管线：{pipeline_text}")
    if total_docs > 0:
        summary_parts.append(f"覆盖文献 {total_docs}")
    return " · ".join(part for part in summary_parts if part)


def _generate_scatter_assets(
    points: List[Dict[str, Any]],
    cluster_count: int,
    cluster_algo: str,
    explained: float,
    pipeline_text: str,
    total_docs: int,
    slug: str,
) -> Dict[str, Any]:
    if not points:
        return {}

    x_values = np.array([float(p.get("x", 0.0)) for p in points], dtype=float)
    y_values = np.array([float(p.get("y", 0.0)) for p in points], dtype=float)
    counts = np.array([max(1, int(p.get("count", 0))) for p in points], dtype=float)
    sqrt_counts = np.sqrt(counts)

    if np.all(np.isfinite(sqrt_counts)) and sqrt_counts.max() > sqrt_counts.min():
        norm = (sqrt_counts - sqrt_counts.min()) / (sqrt_counts.max() - sqrt_counts.min())
    else:
        norm = np.zeros_like(sqrt_counts)

    marker_sizes = 120 + norm * 420
    has_clusters = cluster_count > 1
    colors = [
        _scatter_color_for_cluster(int(p.get("cluster", 0)), has_clusters)
        for p in points
    ]

    fig, ax = plt.subplots(figsize=(7.2, 5.4))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    ax.grid(True, linestyle="--", linewidth=0.6, alpha=0.25)
    ax.set_xlabel("Semantic Dimension 1", fontname="Times New Roman", fontsize=12)
    ax.set_ylabel("Semantic Dimension 2", fontname="Times New Roman", fontsize=12)
    ax.tick_params(axis="both", labelsize=10)
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontname("Times New Roman")

    scatter = ax.scatter(
        x_values,
        y_values,
        s=marker_sizes,
        c=colors,
        alpha=0.88,
        edgecolors="#ffffff",
        linewidths=1.1,
    )

    x_pad = max(1e-9, float(np.nanmax(x_values) - np.nanmin(x_values))) * 0.08
    y_pad = max(1e-9, float(np.nanmax(y_values) - np.nanmin(y_values))) * 0.08
    ax.margins(0.08)
    ax.set_xlim(np.nanmin(x_values) - x_pad, np.nanmax(x_values) + x_pad)
    ax.set_ylim(np.nanmin(y_values) - y_pad, np.nanmax(y_values) + y_pad)

    sorted_idx = np.argsort(-counts)
    top_k = min(12, len(points))
    offset = y_pad if y_pad > 0 else 0.05
    for idx in sorted_idx[:top_k]:
        label = str(points[idx].get("label") or "")
        if not label:
            continue
        text_kwargs: Dict[str, Any] = {
            "fontsize": 10,
            "color": "#0f172a",
            "ha": "center",
            "va": "bottom",
        }
        if _contains_cjk(label) and SCATTER_CHINESE_FONT is not None:
            text_kwargs["fontproperties"] = SCATTER_CHINESE_FONT
        else:
            text_kwargs["fontname"] = "Times New Roman"
        ax.text(x_values[idx], y_values[idx] + offset, label, **text_kwargs)

    summary = _scatter_summary(cluster_count, cluster_algo, explained, pipeline_text, total_docs)
    ax.set_title("语义二维映射", fontname="Times New Roman", fontsize=14, color="#0f172a")
    ax.annotate(
        summary,
        xy=(0, 1.02),
        xycoords="axes fraction",
        ha="left",
        va="bottom",
        fontsize=10,
        fontname="Times New Roman",
        color="#475569",
    )

    legend_elements = []
    if has_clusters:
        legend_elements = []
        unique_clusters = sorted({int(p.get("cluster", 0)) for p in points if int(p.get("cluster", 0)) >= 0})
        for cluster in unique_clusters[:8]:
            legend_elements.append(
                plt.Line2D(
                    [0],
                    [0],
                    marker="o",
                    color="w",
                    markerfacecolor=_scatter_color_for_cluster(cluster, True),
                    markeredgecolor="#ffffff",
                    markersize=8,
                    label=f"簇 {cluster + 1}",
                )
            )
        if any(int(p.get("cluster", 0)) < 0 for p in points):
            legend_elements.append(
                plt.Line2D(
                    [0],
                    [0],
                    marker="o",
                    color="w",
                    markerfacecolor="#ef4444",
                    markeredgecolor="#ffffff",
                    markersize=8,
                    label="噪声",
                )
            )
    if legend_elements:
        ax.legend(handles=legend_elements, loc="upper right", frameon=False, fontsize=9)

    fig.tight_layout()

    png_buffer = io.BytesIO()
    svg_buffer = io.BytesIO()
    fig.savefig(png_buffer, format="png", dpi=300, bbox_inches="tight", facecolor="white")
    fig.savefig(svg_buffer, format="svg", bbox_inches="tight", facecolor="white")
    plt.close(fig)

    png_data = base64.b64encode(png_buffer.getvalue()).decode("ascii")
    svg_data = base64.b64encode(svg_buffer.getvalue()).decode("ascii")
    png_buffer.close()
    svg_buffer.close()

    base = _slugify_filename(slug or "scatter") or "scatter"
    png_filename = f"{base}_semantic_map.png"
    svg_filename = f"{base}_semantic_map.svg"

    return {
        "png_data_url": f"data:image/png;base64,{png_data}",
        "svg_data_url": f"data:image/svg+xml;base64,{svg_data}",
        "png_filename": png_filename,
        "svg_filename": svg_filename,
    }
WORDCLOUD_MAX_WORDS = int(os.getenv("WORDCLOUD_MAX_WORDS", "120"))

def _get_lock(sid: str) -> Lock:
    with SESSION_LOCKS_LOCK:
        if sid not in SESSION_LOCKS:
            SESSION_LOCKS[sid] = Lock()
        return SESSION_LOCKS[sid]

def _touch(sid: str):
    SESSION_TOUCH[sid] = time.time()

def _evict_if_needed():
    """LRU：将超过 MAX_MEMORY_SESSIONS 的最久未用会话从内存卸载（磁盘已有持久化）。"""
    try:
        if len(DATASTORE) <= MAX_MEMORY_SESSIONS:
            return
        # 依据最新访问时间排序，保留最近的
        alive = sorted(SESSION_TOUCH.items(), key=lambda kv: kv[1], reverse=True)
        keep = set([sid for sid, _ in alive[:MAX_MEMORY_SESSIONS]])
        remove = [sid for sid in DATASTORE.keys() if sid not in keep]
        for sid in remove:
            # 不删除磁盘文件，只释放内存
            try:
                del DATASTORE[sid]
                TOPIC_VIZ_CACHE.pop(sid, None)
            except Exception:
                pass
        gc.collect()
    except Exception:
        pass

def _ensure_session_loaded(session_id: str) -> pd.DataFrame:
    """返回指定 session 对应的 DataFrame，如有需要则从磁盘加载。"""

    if not session_id:
        raise ValueError("session_id is required")
    if session_id not in DATASTORE:
        try:
            with _get_lock(session_id):
                DATASTORE[session_id] = load_session_from_disk(session_id)
                _touch(session_id)
                _evict_if_needed()
        except Exception as exc:
            raise ValueError("invalid session_id") from exc
    else:
        _touch(session_id)
        _evict_if_needed()
    return DATASTORE[session_id]


# ================== 参考列表 ==================
TOPIC_LIST = [
    "健康议题","经济议题","政治议题","环境议题","传播模式与行为",
    "媒介制度与平台治理","科技议题","文化议题","宗教议题","其他议题"
]
FIELD_LIST = [
    "广告学",
    "传播心理学",
    "传播研究方法",
    "传播伦理",
    "健康传播",
    "跨文化传播",
    "人际传播",
    "新闻学",
    "法律与政策",
    "大众传播",
    "媒介效果",
    "媒介史",
    "媒介产业",
    "媒介生产与传播",
    "媒介技术",
    "媒介理论",
    "组织传播",
    "政治传播",
    "公共关系",
    "科学传播",
    "言语传播",
    "其他领域"
]

LABEL_MAPPING_LOCK = Lock()
LABEL_MAPPING_FILE = os.path.join(OUTPUT_DIR, "label_mappings.json")

DEFAULT_TOPIC_LABEL_MAP: Dict[str, str] = {
    "文化议题": "Cultural Issues",
    "经济议题": "Economic Issues",
    "环境议题": "Environmental Issues",
    "健康议题": "Health Issues",
    "媒介制度与平台治理": "Media Systems",
    "政治议题": "Political Issues",
    "宗教议题": "Religious Issues",
    "科技议题": "Technological Issues",
    "传播模式与行为": "Communication Behaviors",
    "其他议题": "Other Issues",
    "未分类": "Uncategorized",
}

DEFAULT_FIELD_LABEL_MAP: Dict[str, str] = {
    "广告学": "Advertising",
    "传播心理学": "Communication & Psychology",
    "传播研究方法": "Communication Research Methods",
    "传播伦理": "Communication Ethics",
    "健康传播": "Health Communications",
    "跨文化传播": "Intercultural Communications",
    "人际传播": "Interpersonal Communications",
    "新闻学": "Journalism",
    "法律与政策": "Law & Policy",
    "大众传播": "Mass Communication",
    "媒介效果": "Media Effects",
    "媒介史": "Media History",
    "媒介产业": "Media Industries",
    "媒介生产与传播": "Media Production & Distribution",
    "媒介技术": "Media Technology",
    "媒介理论": "Media Theory",
    "组织传播": "Organizational Communications",
    "政治传播": "Political Communications",
    "公共关系": "Public Relations",
    "科学传播": "Science Communications",
    "言语传播": "Speech Communications",
    "其他领域": "Others",
    "未分类": "Uncategorized",
}

TOPIC_LABEL_OVERRIDES: Dict[str, str] = {}
FIELD_LABEL_OVERRIDES: Dict[str, str] = {}
TOPIC_LABEL_MAP: Dict[str, str] = {}
FIELD_LABEL_MAP: Dict[str, str] = {}

CATEGORY_BAR_CONFIG = {
    "topic_adj": {
        "column": ADJUST_TOPIC_COL,
        "axis": "Topic",
        "title": "Topic Distribution (Adjusted)",
        "kind": "topic",
    },
    "field_adj": {
        "column": ADJUST_FIELD_COL,
        "axis": "Area",
        "title": "Area Distribution (Adjusted)",
        "kind": "field",
    },
    "topic_orig": {
        "column": "研究主题（议题）分类",
        "axis": "Topic",
        "title": "Topic Distribution (Original)",
        "kind": "topic",
    },
    "field_orig": {
        "column": "研究领域分类",
        "axis": "Area",
        "title": "Area Distribution (Original)",
        "kind": "field",
    },
}


def _sanitize_label_mapping_payload(payload: Any) -> Dict[str, str]:
    if not payload:
        return {}
    if isinstance(payload, dict):
        items = payload.items()
    else:
        try:
            items = dict(payload).items()
        except Exception:
            return {}
    result: Dict[str, str] = {}
    for key, value in items:
        if not isinstance(key, str):
            continue
        name = key.strip()
        if not name:
            continue
        text = "" if value is None else str(value)
        result[name] = text.strip()
    return result


def _compose_label_map(overrides: Dict[str, str], defaults: Dict[str, str]) -> Dict[str, str]:
    merged = defaults.copy()
    merged.update(overrides)
    return merged


def _load_label_mapping_file() -> Tuple[Dict[str, str], Dict[str, str]]:
    if not os.path.exists(LABEL_MAPPING_FILE):
        return {}, {}
    try:
        with open(LABEL_MAPPING_FILE, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception:
        return {}, {}
    topic_raw = data.get("topic") if isinstance(data, dict) else {}
    field_raw = data.get("field") if isinstance(data, dict) else {}
    return _sanitize_label_mapping_payload(topic_raw), _sanitize_label_mapping_payload(field_raw)


def _save_label_mapping_file(topic_overrides: Dict[str, str], field_overrides: Dict[str, str]) -> None:
    os.makedirs(os.path.dirname(LABEL_MAPPING_FILE), exist_ok=True)
    with open(LABEL_MAPPING_FILE, "w", encoding="utf-8") as fh:
        json.dump({"topic": topic_overrides, "field": field_overrides}, fh, ensure_ascii=False, indent=2)


def _update_label_mappings(topic_overrides: Dict[str, str], field_overrides: Dict[str, str]) -> None:
    global TOPIC_LABEL_OVERRIDES, FIELD_LABEL_OVERRIDES, TOPIC_LABEL_MAP, FIELD_LABEL_MAP
    TOPIC_LABEL_OVERRIDES = topic_overrides
    FIELD_LABEL_OVERRIDES = field_overrides
    TOPIC_LABEL_MAP = _compose_label_map(TOPIC_LABEL_OVERRIDES, DEFAULT_TOPIC_LABEL_MAP)
    FIELD_LABEL_MAP = _compose_label_map(FIELD_LABEL_OVERRIDES, DEFAULT_FIELD_LABEL_MAP)


def _initialize_label_mappings() -> None:
    topic_overrides, field_overrides = _load_label_mapping_file()
    _update_label_mappings(topic_overrides, field_overrides)


def _label_display(label: str, mapping: Dict[str, str]) -> str:
    if not isinstance(label, str):
        return ""
    key = label.strip()
    if not key:
        return ""
    value = mapping.get(key)
    return value.strip() if isinstance(value, str) else ""


_initialize_label_mappings()
REQUIRED_COLS = [
    "Article Title","Abstract","结构化总结",
    "研究主题（议题）分类","研究领域分类",
    "Publication Year","DOI"
]
# ================== FastAPI 应用 ==================
app = FastAPI(title="分类人工调整 GUI", version=APP_VERSION)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/", response_class=HTMLResponse)
def index():
    """Serve the main HTML shell with an explicit UTF-8 content-type."""
    index_path = os.path.join(STATIC_DIR, "index.html")
    try:
        with open(index_path, "r", encoding="utf-8") as fh:
            return HTMLResponse(content=fh.read(), media_type="text/html; charset=utf-8")
    except FileNotFoundError:
        return HTMLResponse(
            content="<h1>index.html not found</h1>",
            status_code=404,
            media_type="text/html; charset=utf-8",
        )

# ================== 进度条 ==================
def _progress_init(sid:str):
    with PROGRESS_LOCK:
        PROGRESS[sid] = {"status":"idle","message":"","percent":0,"emb_current":0,"emb_total":0,"groups_done":0,"groups_total":0,"outliers":0}

def _progress_update(sid:str, **kw):
    with PROGRESS_LOCK:
        st = PROGRESS.setdefault(sid, {})
        st.update(kw)
        emb_total = st.get("emb_total",0) or 0
        emb_cur   = st.get("emb_current",0) or 0
        g_total   = st.get("groups_total",0) or 0
        g_done    = st.get("groups_done",0) or 0
        emb_prog  = (emb_cur/emb_total) if emb_total else 0
        grp_prog  = (g_done/g_total) if g_total else 0
        st["percent"] = int(min(100, round(emb_prog*70 + grp_prog*30)))
        PROGRESS[sid] = st

def _progress_finish(sid:str, ok=True, msg=""):
    with PROGRESS_LOCK:
        st = PROGRESS.setdefault(sid,{})
        st["status"] = "done" if ok else "error"
        st["message"]= msg
        if ok: st["percent"]=100
        PROGRESS[sid] = st

@app.get("/progress")
def progress(session_id: str = Query(...)):
    with PROGRESS_LOCK:
        st = PROGRESS.get(session_id, {"status":"idle","percent":0,"message":""})
    return JSONResponse(st)

# ================== 工具函数 ==================
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    for c in REQUIRED_COLS:
        if c not in df.columns: df[c] = ""
    if ADJUST_TOPIC_COL not in df.columns: df[ADJUST_TOPIC_COL] = ""
    if ADJUST_FIELD_COL not in df.columns: df[ADJUST_FIELD_COL] = ""
    for c in [RANK_SCORE_COL,RANK_GROUP_COL,RANK_OUTLIER_COL,RANK_ALGO_COL]:
        if c not in df.columns: df[c] = "" if c != RANK_OUTLIER_COL else False
    # 统一类型，避免写盘/读盘后类型漂移
    df[ADJUST_TOPIC_COL] = df[ADJUST_TOPIC_COL].astype(str).replace("nan","")
    df[ADJUST_FIELD_COL] = df[ADJUST_FIELD_COL].astype(str).replace("nan","")
    if RANK_OUTLIER_COL in df.columns:
        df[RANK_OUTLIER_COL] = df[RANK_OUTLIER_COL].astype(bool)
    return df

def read_table(file: UploadFile) -> pd.DataFrame:
    name = (file.filename or "").lower()
    bio = io.BytesIO(file.file.read())
    if name.endswith((".xlsx",".xls")):
        df = pd.read_excel(bio)
    elif name.endswith(".csv"):
        try: df = pd.read_csv(bio, encoding="utf-8-sig")
        except Exception:
            bio.seek(0); df = pd.read_csv(bio, encoding="gb18030")
    else:
        raise ValueError("仅支持 .xlsx/.xls/.csv")
    return normalize_columns(df)

def safe_get(row: pd.Series, col: str) -> str:
    v = row.get(col, "")
    return "" if pd.isna(v) else str(v)

def _counts(ser: pd.Series, mapping: Optional[Dict[str, str]] = None) -> List[Dict[str,Any]]:
    s = ser.fillna("").astype(str).str.strip()
    s = s[s!=""]
    total = int(s.shape[0]) if s.shape[0] else 1
    vc = s.value_counts()
    result: List[Dict[str, Any]] = []
    mapping = mapping or {}
    for k, v in vc.items():
        label = str(k)
        item = {
            "label": label,
            "count": int(v),
            "percent": round(v * 100.0 / total, 2),
        }
        item["display_label"] = _label_display(label, mapping)
        result.append(item)
    return result


def _invalidate_topic_viz(sid: str):
    try:
        TOPIC_VIZ_CACHE.pop(sid, None)
    except Exception:
        pass


def _tokenize_text(text: str) -> List[str]:
    if not text:
        return []
    raw = TOKEN_PATTERN.findall(str(text))
    tokens: List[str] = []
    for tok in raw:
        if not tok:
            continue
        if re.fullmatch(r"[A-Za-z]{2,}", tok):
            norm = tok.lower()
        else:
            norm = tok
        if norm in STOPWORDS:
            continue
        tokens.append(norm)
    return tokens


def _combine_text_columns(df: pd.DataFrame, columns: List[str]) -> List[str]:
    cols = [c for c in columns if c in df.columns]
    if not cols:
        return []
    series = df[cols[0]].fillna("").astype(str)
    for c in cols[1:]:
        series = series + " " + df[c].fillna("").astype(str)
    return series.tolist()


def _category_keywords(
    df: pd.DataFrame,
    column: str,
    fallback_label: str,
    top_n: int = 80,
    label_mapping: Optional[Dict[str, str]] = None,
) -> Tuple[List[Dict[str, Any]], List[str], List[int], List[str]]:
    if column not in df.columns or df.empty:
        return [], [], [], []
    text_columns = [c for c in KEYWORD_TEXT_COLUMNS if c in df.columns]
    work_columns = [column] + text_columns
    work = df[work_columns].copy()
    work[column] = work[column].fillna("").astype(str).str.strip().replace("", fallback_label)
    categories: List[Dict[str, Any]] = []
    texts: List[str] = []
    labels: List[str] = []
    counts: List[int] = []
    mapping = label_mapping or {}
    for label, group in work.groupby(column):
        label_str = str(label).strip() or fallback_label
        total = int(group.shape[0])
        combined_texts = _combine_text_columns(group, text_columns)
        freq: Counter[str] = Counter()
        for text in combined_texts:
            freq.update(_tokenize_text(text))
        total_tokens = sum(freq.values())
        top_words = [
            {
                "token": tok,
                "count": int(cnt),
                "weight": round(cnt / total_tokens, 4) if total_tokens else 0.0,
            }
            for tok, cnt in freq.most_common(top_n)
        ]
        categories.append(
            {
                "label": label_str,
                "count": total,
                "top_words": top_words,
                "total_tokens": int(total_tokens),
                "display_label": _label_display(label_str, mapping),
            }
        )
        labels.append(label_str)
        counts.append(total)
        texts.append(" ".join(combined_texts))
    categories.sort(key=lambda x: (-x.get("count", 0), x.get("label", "")))
    return categories, texts, counts, labels


def _category_scatter(
    texts: List[str],
    labels: List[str],
    counts: List[int],
    slug: str,
    label_mapping: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    n = len(labels)
    if n == 0:
        return {"points": [], "cluster_count": 0, "explained_variance": 0.0}

    vectorizer_name = ""
    reduction_method = "pca"
    arr: Optional[np.ndarray] = None

    if n:
        try:
            arr = embed_local(texts, None)
            if arr.shape[0] == n:
                vectorizer_name = "sbert"
            else:
                arr = None
        except Exception:
            arr = None

    if arr is None:
        vectorizer_name = "tfidf"
        try:
            vectorizer = TfidfVectorizer(
                tokenizer=_tokenize_text,
                lowercase=True,
                token_pattern=None,
                max_features=2000,
            )
            matrix = vectorizer.fit_transform(texts)
            arr = matrix.toarray()
        except Exception:
            arr = np.zeros((n, 1), dtype=float)

    if arr.shape[0] != n:
        arr = np.zeros((n, max(1, arr.shape[1] if arr.ndim == 2 else 1)), dtype=float)

    coords = np.zeros((n, 2), dtype=float)
    explained = 0.0
    if n == 1:
        coords[0, :] = 0.0
    else:
        try:
            if arr.shape[1] >= 2:
                pca = PCA(n_components=2)
                coords = pca.fit_transform(arr)
                explained = float(np.sum(pca.explained_variance_ratio_))
            elif arr.shape[1] == 1:
                coords[:, 0] = arr[:, 0]
            else:
                coords = np.zeros((n, 2), dtype=float)
        except Exception:
            coords = np.zeros((n, 2), dtype=float)

    cluster_labels = np.full(n, -1, dtype=int)
    cluster_count = 0
    cluster_algo: Optional[str] = None

    if n >= 2 and hdbscan is not None:
        try:
            min_cluster_size = max(2, int(math.sqrt(n)))
            clusterer = hdbscan.HDBSCAN(min_cluster_size=min_cluster_size, metric="euclidean")
            raw_labels = clusterer.fit_predict(arr)
            probs = getattr(clusterer, "probabilities_", None)

            valid_labels = sorted({int(lbl) for lbl in raw_labels.tolist() if lbl >= 0})
            label_map = {lbl: idx for idx, lbl in enumerate(valid_labels)}
            has_noise = bool(any(int(l) < 0 for l in raw_labels.tolist()))

            for i, lbl in enumerate(raw_labels):
                if lbl >= 0:
                    cluster_labels[i] = label_map[int(lbl)]
                else:
                    cluster_labels[i] = len(valid_labels)

            cluster_count = len(valid_labels)
            if has_noise and valid_labels:
                cluster_count += 1
            if not valid_labels:
                cluster_labels[:] = 0
                cluster_count = 1 if n > 0 else 0
            cluster_algo = "hdbscan"

            if probs is not None and probs.shape[0] == n:
                explained = float(max(explained, 1.0 - float(np.mean(probs))))
        except Exception:
            cluster_labels = np.zeros(n, dtype=int)
            cluster_count = 0
            cluster_algo = None

    mapping = label_mapping or {}
    points = []
    for i in range(n):
        points.append(
            {
                "label": labels[i],
                "x": float(coords[i, 0]),
                "y": float(coords[i, 1]),
                "count": int(counts[i]),
                "cluster": int(cluster_labels[i]) if cluster_count else 0,
                "display_label": _label_display(labels[i], mapping),
            }
        )

    if cluster_count <= 0:
        effective_clusters = 1 if n else 0
        unique_clusters = effective_clusters
    else:
        unique_clusters = cluster_count
        effective_clusters = cluster_count

    explained = max(0.0, min(1.0, float(explained)))
    total_docs = int(sum(max(0, c) for c in counts))
    pipeline_text = _format_pipeline(vectorizer_name, reduction_method)
    summary_text = _scatter_summary(effective_clusters, cluster_algo or "", explained, pipeline_text, total_docs)

    assets: Dict[str, Any] = {}
    try:
        assets = _generate_scatter_assets(
            points,
            effective_clusters,
            cluster_algo or "",
            explained,
            pipeline_text,
            total_docs,
            slug,
        )
    except Exception:
        traceback.print_exc()
        assets = {}

    return {
        "points": points,
        "cluster_count": int(effective_clusters),
        "unique_clusters": int(unique_clusters),
        "explained_variance": round(explained, 4),
        "cluster_algorithm": cluster_algo or "",
        "vectorizer": vectorizer_name,
        "reduction": reduction_method,
        "document_total": total_docs,
        "pipeline_text": pipeline_text,
        "summary": summary_text,
        **assets,
    }


def _prepare_keyword_dashboard(
    categories: List[Dict[str, Any]],
    *,
    label_mapping: Optional[Dict[str, str]] = None,
    limit: int = 15,
) -> List[Dict[str, Any]]:
    if not categories:
        return []

    dashboard: List[Dict[str, Any]] = []
    mapping = label_mapping or {}

    for category in categories:
        if not isinstance(category, dict):
            continue

        words = category.get("top_words") or []
        if not isinstance(words, list):
            continue

        keywords: List[str] = []
        weights: List[float] = []
        counts: List[int] = []

        for word in words:
            if not isinstance(word, dict):
                continue
            token = str(word.get("token", "")).strip()
            if not token:
                continue
            weight_raw = word.get("weight", 0.0)
            try:
                weight = float(weight_raw or 0.0)
            except (TypeError, ValueError):
                weight = 0.0
            count_raw = word.get("count", 0)
            try:
                count = int(count_raw or 0)
            except (TypeError, ValueError):
                count = 0

            keywords.append(token)
            weights.append(round(weight, 4))
            counts.append(count)

            if len(keywords) >= limit:
                break

        if not keywords:
            continue

        label = str(category.get("label", "") or "未分类")
        display = str(
            category.get("display_label")
            or _label_display(label, mapping)
            or label
        )

        dashboard.append(
            {
                "label": label,
                "display_label": display,
                "document_count": int(category.get("count", 0) or 0),
                "total_tokens": int(category.get("total_tokens", 0) or 0),
                "keywords": keywords,
                "weights": weights,
                "counts": counts,
            }
        )

    return dashboard


def _topic_visual_payload(df: pd.DataFrame) -> Dict[str, Any]:
    topic_categories, topic_texts, topic_counts, topic_labels = _category_keywords(
        df, ADJUST_TOPIC_COL, "未分类", label_mapping=TOPIC_LABEL_MAP
    )
    field_categories, field_texts, field_counts, field_labels = _category_keywords(
        df, ADJUST_FIELD_COL, "未分类", label_mapping=FIELD_LABEL_MAP
    )
    topic_scatter = _category_scatter(
        topic_texts, topic_labels, topic_counts, "topic_adj", label_mapping=TOPIC_LABEL_MAP
    )
    field_scatter = _category_scatter(
        field_texts, field_labels, field_counts, "field_adj", label_mapping=FIELD_LABEL_MAP
    )
    topic_dashboard = _prepare_keyword_dashboard(
        topic_categories, label_mapping=TOPIC_LABEL_MAP
    )
    field_dashboard = _prepare_keyword_dashboard(
        field_categories, label_mapping=FIELD_LABEL_MAP
    )

    return {
        "topic_adj": {
            "categories": topic_categories,
            "scatter": topic_scatter,
            "keyword_dashboard": topic_dashboard,
        },
        "field_adj": {
            "categories": field_categories,
            "scatter": field_scatter,
            "keyword_dashboard": field_dashboard,
        },
    }


def _get_topic_visual_data(session_id: str) -> Dict[str, Any]:
    df = _ensure_session_loaded(session_id)
    cache = TOPIC_VIZ_CACHE.get(session_id)
    payload: Dict[str, Any] = {}
    if cache:
        payload = cache.get("data", {}) or {}
    if not payload:
        payload = _topic_visual_payload(df)
        cache = {"data": payload, "updated_at": time.time()}
        TOPIC_VIZ_CACHE[session_id] = cache
    else:
        cache["updated_at"] = time.time()
    return payload


def _slugify_filename(text: str) -> str:
    base = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", text).strip("_")
    return base or "wordcloud"


def _generate_word_cloud_assets(words: List[Dict[str, Any]]) -> Dict[str, Any]:
    frequencies = {}
    for item in words:
        token = str(item.get("token", "")).strip()
        count = int(item.get("count", 0))
        if not token or count <= 0:
            continue
        frequencies[token] = count
        if len(frequencies) >= WORDCLOUD_MAX_WORDS:
            break
    if not frequencies:
        raise ValueError("no frequencies")
    wc = WordCloud(
        width=WORDCLOUD_WIDTH,
        height=WORDCLOUD_HEIGHT,
        background_color="white",
        font_path=WORDCLOUD_FONT_PATH,
        prefer_horizontal=0.95,
        random_state=42,
        max_words=WORDCLOUD_MAX_WORDS,
        collocations=False,
        margin=2,
    )
    wc.generate_from_frequencies(frequencies)

    png_buffer = io.BytesIO()
    image = wc.to_image()
    image.save(png_buffer, format="PNG", dpi=(300, 300))
    png_data = base64.b64encode(png_buffer.getvalue()).decode("ascii")

    try:
        svg_text = wc.to_svg(embed_font=bool(WORDCLOUD_FONT_PATH))
    except Exception:
        svg_text = wc.to_svg(embed_font=False)
    svg_data = base64.b64encode(svg_text.encode("utf-8")).decode("ascii")

    return {
        "png_data_url": f"data:image/png;base64,{png_data}",
        "svg_data_url": f"data:image/svg+xml;base64,{svg_data}",
        "width": wc.width,
        "height": wc.height,
        "words_used": len(frequencies),
    }


def _generate_keyword_bar_assets(
    words: List[Dict[str, Any]],
    top_n: int = 12,
) -> Dict[str, Any]:
    entries: List[Tuple[str, int, float]] = []
    for item in words:
        token = str(item.get("token", "")).strip()
        count = int(item.get("count", 0) or 0)
        weight = float(item.get("weight", 0.0) or 0.0)
        if not token or count <= 0:
            continue
        entries.append((token, count, weight))
        if len(entries) >= top_n:
            break
    if not entries:
        raise ValueError("no keyword frequencies")

    labels = [token for token, _, _ in entries]
    counts = [count for _, count, _ in entries]
    weights = [weight for _, _, weight in entries]

    figure_height = max(2.8, len(labels) * 0.5 + 1.6)
    figure_width = max(4.2, 5.2)
    max_dim = max(figure_height, figure_width)
    if max_dim > 8.0:
        scale = 8.0 / max_dim
        figure_height *= scale
        figure_width *= scale

    fig, ax = plt.subplots(figsize=(figure_width, figure_height))
    positions = np.arange(len(labels))
    bars = ax.barh(positions, counts, color="#2563eb", alpha=0.9)

    ax.set_yticks(positions)
    ax.set_yticklabels(labels, fontsize=11, fontfamily="Times New Roman")
    ax.set_xlabel("Occurrence Count", fontsize=12)
    ax.set_ylabel("Keyword", fontsize=12)
    ax.invert_yaxis()

    ax.set_axisbelow(True)
    ax.grid(False)
    ax.set_facecolor("#ffffff")
    fig.patch.set_facecolor("#ffffff")

    for spine_name, spine in ax.spines.items():
        if spine_name in ("left", "bottom"):
            spine.set_visible(True)
            spine.set_color("#1f2937")
            spine.set_linewidth(1.1)
        else:
            spine.set_visible(False)

    ax.tick_params(axis="x", labelsize=11, direction="out", length=4, width=0.8)
    ax.tick_params(axis="y", labelsize=11, direction="out", length=4, width=0.8)

    bar_labels = []
    for idx, value in enumerate(counts):
        percent = weights[idx] * 100 if idx < len(weights) else 0.0
        if percent > 0:
            bar_labels.append(f"{value:,} ({percent:.1f}%)")
        else:
            bar_labels.append(f"{value:,}")
    ax.bar_label(bars, labels=bar_labels, padding=6, fontsize=10, color="#1f2937")

    fig.tight_layout()

    png_buffer = io.BytesIO()
    svg_buffer = io.BytesIO()
    fig.savefig(
        png_buffer,
        format="png",
        dpi=300,
        bbox_inches="tight",
        facecolor="white",
    )
    fig.savefig(svg_buffer, format="svg", bbox_inches="tight", facecolor="white")
    plt.close(fig)

    png_data = base64.b64encode(png_buffer.getvalue()).decode("ascii")
    svg_data = base64.b64encode(svg_buffer.getvalue()).decode("ascii")

    return {
        "png_data_url": f"data:image/png;base64,{png_data}",
        "svg_data_url": f"data:image/svg+xml;base64,{svg_data}",
        "words_used": len(labels),
    }


def _generate_category_bar_assets(
    entries: List[Dict[str, Any]],
    *,
    dimension: str,
    axis_label: str,
    title: str,
) -> Dict[str, Any]:
    filtered: List[Tuple[str, str, int]] = []
    for item in entries:
        if not isinstance(item, dict):
            continue
        try:
            count = int(item.get("count", 0) or 0)
        except (TypeError, ValueError):
            count = 0
        if count <= 0:
            continue
        label = str(item.get("label", "")).strip()
        display = str(item.get("display_label") or label or "").strip()
        if not display:
            display = label or "Unlabeled"
        filtered.append((label or display, display, count))

    if not filtered:
        raise ValueError("no category counts")

    filtered.sort(key=lambda row: (-row[2], row[1]))

    labels = [display for _, display, _ in filtered]
    counts = [count for _, _, count in filtered]

    width = 6.2
    height = max(3.2, 0.45 * len(labels) + 1.6)
    max_dim = max(width, height)
    if max_dim > 8.0:
        scale = 8.0 / max_dim
        width *= scale
        height *= scale

    cmap = matplotlib.colormaps.get("viridis")
    if cmap is None:
        colors = ["#0f172a" for _ in labels]
    else:
        colors = [cmap(0.2 + 0.6 * (idx / max(1, len(labels) - 1))) for idx in range(len(labels))]

    fig, ax = plt.subplots(figsize=(width, height))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    positions = np.arange(len(labels))
    bars = ax.barh(positions, counts, color=colors, alpha=0.9, edgecolor="white", linewidth=1.0)

    ax.set_yticks(positions)
    ax.set_yticklabels(labels, fontname="Times New Roman", fontsize=11)
    ax.set_xlabel("Document Count", fontname="Times New Roman", fontsize=12)
    ax.set_ylabel(axis_label, fontname="Times New Roman", fontsize=12)
    ax.set_title(title, fontname="Times New Roman", fontsize=14, color="#111827", pad=14)
    ax.invert_yaxis()

    ax.set_axisbelow(True)
    ax.grid(False)

    for spine_name, spine in ax.spines.items():
        if spine_name in ("left", "bottom"):
            spine.set_visible(True)
            spine.set_color("#111827")
            spine.set_linewidth(1.1)
        else:
            spine.set_visible(False)

    ax.tick_params(axis="x", labelsize=11, direction="out", length=4, width=0.8, colors="#111827")
    ax.tick_params(axis="y", labelsize=11, direction="out", length=4, width=0.8, colors="#111827")
    for label in ax.get_xticklabels():
        label.set_fontname("Times New Roman")

    ax.bar_label(
        bars,
        labels=[f"{value:,}" for value in counts],
        padding=6,
        fontsize=10,
        color="#111827",
        fontname="Times New Roman",
    )

    fig.tight_layout()

    png_buffer = io.BytesIO()
    svg_buffer = io.BytesIO()
    fig.savefig(
        png_buffer,
        format="png",
        dpi=300,
        bbox_inches="tight",
        facecolor="white",
    )
    fig.savefig(
        svg_buffer,
        format="svg",
        bbox_inches="tight",
        facecolor="white",
    )
    plt.close(fig)

    png_data = base64.b64encode(png_buffer.getvalue()).decode("ascii")
    svg_data = base64.b64encode(svg_buffer.getvalue()).decode("ascii")
    png_buffer.close()
    svg_buffer.close()

    slug = _slugify_filename(f"{dimension}_distribution") or "distribution"

    return {
        "png_data_url": f"data:image/png;base64,{png_data}",
        "svg_data_url": f"data:image/svg+xml;base64,{svg_data}",
        "png_filename": f"{slug}.png",
        "svg_filename": f"{slug}.svg",
        "title": title,
    }


def _top_sources_by_category(
    df: pd.DataFrame,
    category_col: str,
    source_col: str = "Source Title",
    fallback_category: str = "未分类",
    fallback_source: str = "未提供期刊",
    top_n: int = 5,
    max_categories: int = 24,
    label_mapping: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    """统计每个分类下按 Source Title 排名前 N 的期刊列表。"""

    if source_col not in df.columns or category_col not in df.columns:
        return []

    work_df = df[[category_col, source_col]].copy()
    work_df[category_col] = (
        work_df[category_col]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", fallback_category)
    )
    work_df[source_col] = (
        work_df[source_col]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", fallback_source)
    )

    if work_df.empty:
        return []

    grouped = (
        work_df.groupby([category_col, source_col]).size().reset_index(name="count")
    )
    if grouped.empty:
        return []

    totals = grouped.groupby(category_col)["count"].sum()
    if totals.empty:
        return []

    ordered_categories = (
        totals.sort_values(ascending=False).index.tolist()
    )
    if max_categories > 0:
        ordered_categories = ordered_categories[:max_categories]

    mapping = label_mapping or {}
    result: List[Dict[str, Any]] = []
    for cat in ordered_categories:
        cat_total = int(totals.get(cat, 0))
        if cat_total <= 0:
            continue
        sub_df = grouped[grouped[category_col] == cat]
        if sub_df.empty:
            continue
        sub_sorted = sub_df.sort_values(
            by=["count", source_col], ascending=[False, True]
        ).head(top_n)
        items: List[Dict[str, Any]] = []
        for _, row in sub_sorted.iterrows():
            count = int(row["count"])
            if count <= 0:
                continue
            items.append(
                {
                    "label": str(row[source_col]),
                    "count": count,
                    "percent": round(count * 100.0 / cat_total, 2) if cat_total else 0.0,
                }
            )
        if not items:
            continue
        result.append(
            {
                "category": str(cat),
                "category_display": _label_display(str(cat), mapping),
                "total": cat_total,
                "unique_sources": int(sub_df.shape[0]),
                "items": items,
            }
        )

    return result


def _quartile_rank(label: str) -> int:
    match = re.search(r"Q\s*([1-4])", str(label), re.IGNORECASE)
    if match:
        try:
            return int(match.group(1))
        except (TypeError, ValueError):
            return 99
    return 99


def _normalise_quartile_label(label: str) -> str:
    text = str(label or "").strip()
    if not text:
        return ""
    match = re.search(r"Q\s*([1-4])", text, re.IGNORECASE)
    if match:
        return f"Q{match.group(1)}"
    return text.upper()


def _journal_overview_table(
    df: pd.DataFrame,
    source_col: str = "Source Title",
    limit: int = 48,
    *,
    citation_candidates: Optional[List[str]] = None,
    quartile_candidates: Optional[List[str]] = None,
    year_candidates: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Aggregate journal metrics for the overview table shown on the dashboard."""

    # Determine which columns are available in the dataset.
    source_candidates = [
        source_col,
        "Source Title",
        "Journal",
        "Journal Title",
    ]
    source_candidates = [col for col in source_candidates if col]
    source_name = next((col for col in source_candidates if col in df.columns), None)
    if not source_name:
        return {
            "rows": [],
            "summary": {
                "listed_journals": 0,
                "unique_journals": 0,
                "total_citations": 0,
                "total_publications": 0,
                "year_min": None,
                "year_max": None,
                "year_column": None,
                "source_column": None,
                "citations_column": None,
                "quartile_column": None,
            },
        }

    citation_candidates = list(
        dict.fromkeys(
            (citation_candidates or [])
            + [
                "Times Cited, All Databases",
                "Times Cited",
                "Times Cited, All Databases ",
            ]
        )
    )
    quartile_candidates = list(
        dict.fromkeys(
            (quartile_candidates or [])
            + [
                "JIF Quartile",
                "JCR Quartile",
                "Journal Quartile",
            ]
        )
    )
    year_candidates = list(
        dict.fromkeys(
            (year_candidates or [])
            + [
                "Publication Year",
                "Year",
                "Year Published",
            ]
        )
    )

    citation_name = next((col for col in citation_candidates if col in df.columns), None)
    quartile_name = next((col for col in quartile_candidates if col in df.columns), None)
    year_name = next((col for col in year_candidates if col in df.columns), None)

    columns = [source_name]
    if citation_name:
        columns.append(citation_name)
    if quartile_name:
        columns.append(quartile_name)
    if year_name:
        columns.append(year_name)

    work = df[columns].copy()
    work.rename(
        columns={
            source_name: "source",
            **({citation_name: "citations"} if citation_name else {}),
            **({quartile_name: "quartile"} if quartile_name else {}),
            **({year_name: "year"} if year_name else {}),
        },
        inplace=True,
    )

    work["source"] = work["source"].fillna("").astype(str).str.strip()
    work = work[work["source"] != ""]
    if work.empty:
        return {
            "rows": [],
            "summary": {
                "listed_journals": 0,
                "unique_journals": 0,
                "total_citations": 0,
                "total_publications": 0,
                "year_min": None,
                "year_max": None,
                "year_column": year_name,
                "source_column": source_name,
                "citations_column": citation_name,
                "quartile_column": quartile_name,
            },
        }

    if "citations" not in work.columns:
        work["citations"] = 0.0
    else:
        work["citations"] = (
            pd.to_numeric(work["citations"], errors="coerce").fillna(0.0)
        )

    if "quartile" not in work.columns:
        work["quartile"] = ""
    else:
        work["quartile"] = work["quartile"].fillna("").astype(str).str.strip()

    year_min: Optional[int] = None
    year_max: Optional[int] = None
    if "year" in work.columns:
        years = pd.to_numeric(work["year"], errors="coerce")
        valid_years = years.dropna()
        if not valid_years.empty:
            year_min = int(valid_years.min())
            year_max = int(valid_years.max())

    grouped = work.groupby("source", sort=False)
    counts = grouped.size().astype(int)
    citations_sum = grouped["citations"].sum()

    def _resolve_quartile(series: pd.Series) -> str:
        values = [
            _normalise_quartile_label(val)
            for val in series
            if str(val or "").strip()
        ]
        if not values:
            return "—"
        occurrences = Counter(values)
        best_label, _ = min(
            occurrences.items(),
            key=lambda item: (
                -item[1],
                _quartile_rank(item[0]),
                item[0],
            ),
        )
        return best_label or "—"

    quartiles = grouped["quartile"].apply(_resolve_quartile)

    records: List[Dict[str, Any]] = []
    for source, publication_count in counts.items():
        citation_value = citations_sum.get(source, 0.0)
        try:
            citations_int = int(round(float(citation_value)))
        except (TypeError, ValueError):
            citations_int = 0
        quartile_value = quartiles.get(source, "—") or "—"
        records.append(
            {
                "name": str(source),
                "citations": citations_int,
                "quartile": quartile_value,
                "publications": int(publication_count),
            }
        )

    records.sort(
        key=lambda item: (
            -item["citations"],
            -item["publications"],
            item["name"].lower(),
        )
    )

    limit = max(int(limit or 0), 0)
    listed = records[:limit] if limit else records

    total_citations = int(sum(item["citations"] for item in listed))
    total_publications = int(sum(item["publications"] for item in listed))

    summary = {
        "listed_journals": len(listed),
        "unique_journals": len(records),
        "total_citations": total_citations,
        "total_publications": total_publications,
        "year_min": year_min,
        "year_max": year_max,
        "year_column": year_name,
        "source_column": source_name,
        "citations_column": citation_name,
        "quartile_column": quartile_name,
    }

    return {"rows": listed, "summary": summary}


def _sanitize_blacklist(raw, allowed: List[str]) -> List[str]:
    """清理来自前端的黑名单，过滤非法值并保持顺序。"""
    if not raw:
        return []
    if isinstance(raw, str):
        candidates = [raw]
    else:
        try:
            candidates = list(raw)
        except TypeError:
            candidates = []
    seen = set()
    cleaned: List[str] = []
    for item in candidates:
        if not isinstance(item, str):
            continue
        name = item.strip()
        if not name or name not in allowed or name in seen:
            continue
        cleaned.append(name)
        seen.add(name)
    return cleaned

def slice_df(df: pd.DataFrame, page:int, page_size:int) -> pd.DataFrame:
    start = max(0,(page-1)*page_size); end = start + page_size
    return df.iloc[start:end].copy()

def _now_str():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _session_dir(sid:str) -> str:
    d = os.path.join(SESS_DIR, sid)
    os.makedirs(d, exist_ok=True)
    return d

def _session_meta_path(sid:str) -> str:
    return os.path.join(_session_dir(sid), "meta.json")

def _session_data_path(sid:str) -> str:
    # 使用 pickle 避免额外依赖（pyarrow/feather）
    return os.path.join(_session_dir(sid), "data.pkl")

def _read_meta(sid:str) -> Dict[str,Any]:
    p = _session_meta_path(sid)
    if not os.path.exists(p): return {}
    try:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _write_meta(sid:str, meta:Dict[str,Any]):
    p = _session_meta_path(sid)
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("[meta write failed]", e)

def save_session_to_disk(sid:str, origin_filename:str=None) -> Dict[str,Any]:
    """将当前会话的 DataFrame + 元信息写盘"""
    if sid not in DATASTORE:
        raise RuntimeError("会话不存在")
    df = DATASTORE[sid]
    dpath = _session_data_path(sid)
    df.to_pickle(dpath)
    meta = _read_meta(sid)
    meta.update({
        "session_id": sid,
        "rows": int(df.shape[0]),
        "updated_at": time.time(),
        "updated_text": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "origin_filename": origin_filename or meta.get("origin_filename",""),
        "app_version": APP_VERSION,
    })
    _write_meta(sid, meta)
    return meta

def load_session_from_disk(sid:str) -> pd.DataFrame:
    p = _session_data_path(sid)
    if not os.path.exists(p):
        raise RuntimeError("会话数据不存在或未保存")
    df = pd.read_pickle(p)
    return normalize_columns(df)

def list_sessions() -> List[Dict[str,Any]]:
    out=[]
    for sid in os.listdir(SESS_DIR):
        sp = os.path.join(SESS_DIR, sid)
        if not os.path.isdir(sp): continue
        meta = _read_meta(sid)
        if not meta:
            # 尝试补充 rows
            dpath = _session_data_path(sid)
            rows = 0
            if os.path.exists(dpath):
                try: rows = pd.read_pickle(dpath).shape[0]
                except: rows = 0
            meta={"session_id":sid,"rows":rows,"updated_at":0}
        out.append({
            "session_id": sid,
            "rows": meta.get("rows",0),
            "updated_text": meta.get("updated_text",""),
            "origin_filename": meta.get("origin_filename",""),
            "last_export_path": meta.get("last_export_path",""),
        })
    # 依据更新时间文本近似排序（无则靠文件ctime）
    def _key(m):
        txt = m.get("updated_text","")
        try:
            return datetime.strptime(txt, "%Y-%m-%d %H:%M:%S")
        except:
            return datetime.fromtimestamp(os.path.getctime(_session_dir(m["session_id"])))
    out.sort(key=_key, reverse=True)
    return out

def _sampling_dir(sid: str) -> str:
    d = os.path.join(_session_dir(sid), SAMPLING_DIR_NAME)
    os.makedirs(d, exist_ok=True)
    return d

def _sampling_slug(inspector: str) -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "-", inspector.strip()).strip("-")
    if not base:
        base = "inspector"
    digest = hashlib.md5(inspector.strip().encode("utf-8")).hexdigest()[:8]
    return f"{base}-{digest}"

def _sampling_path(sid: str, slug: str) -> str:
    return os.path.join(_sampling_dir(sid), f"{slug}.json")

def _read_sampling_audit(sid: str, slug: str) -> Optional[Dict[str, Any]]:
    path = _sampling_path(sid, slug)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _write_sampling_audit(sid: str, slug: str, data: Dict[str, Any]):
    path = _sampling_path(sid, slug)
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def _sampling_item_status(item: Dict[str, Any]) -> str:
    topic = (item.get("selected_topic") or "").strip()
    field = (item.get("selected_field") or "").strip()
    if topic and field:
        return "completed"
    if topic or field:
        return "partial"
    return "pending"

def _sampling_metrics(audit: Dict[str, Any]) -> Dict[str, Any]:
    items = audit.get("items") or []

    def _norm(value: Any) -> str:
        return _normalize_cell_value(value)

    topic_counter: Counter[str] = Counter()
    field_counter: Counter[str] = Counter()
    ai_topic_counter: Counter[str] = Counter()
    ai_field_counter: Counter[str] = Counter()

    topic_compared = 0
    topic_matches = 0
    field_compared = 0
    field_matches = 0
    joint_compared = 0
    joint_matches = 0
    orig_topic_compared = 0
    orig_topic_matches = 0
    orig_field_compared = 0
    orig_field_matches = 0
    orig_joint_compared = 0
    orig_joint_matches = 0

    topic_selected_compared: Counter[str] = Counter()
    topic_ai_compared: Counter[str] = Counter()
    field_selected_compared: Counter[str] = Counter()
    field_ai_compared: Counter[str] = Counter()
    orig_topic_selected_compared: Counter[str] = Counter()
    orig_topic_ref_compared: Counter[str] = Counter()
    orig_field_selected_compared: Counter[str] = Counter()
    orig_field_ref_compared: Counter[str] = Counter()

    for item in items:
        selected_topic = _norm(item.get("selected_topic"))
        selected_field = _norm(item.get("selected_field"))
        ai_topic = _norm(item.get("ai_topic"))
        ai_field = _norm(item.get("ai_field"))
        orig_topic = _norm(item.get("orig_topic"))
        orig_field = _norm(item.get("orig_field"))

        if selected_topic:
            topic_counter[selected_topic] += 1
        if selected_field:
            field_counter[selected_field] += 1
        if ai_topic:
            ai_topic_counter[ai_topic] += 1
        if ai_field:
            ai_field_counter[ai_field] += 1

        if selected_topic and ai_topic:
            topic_compared += 1
            if selected_topic == ai_topic:
                topic_matches += 1
            topic_selected_compared[selected_topic] += 1
            topic_ai_compared[ai_topic] += 1

        if selected_field and ai_field:
            field_compared += 1
            if selected_field == ai_field:
                field_matches += 1
            field_selected_compared[selected_field] += 1
            field_ai_compared[ai_field] += 1

        if selected_topic and selected_field and ai_topic and ai_field:
            joint_compared += 1
            if selected_topic == ai_topic and selected_field == ai_field:
                joint_matches += 1

        if selected_topic and orig_topic:
            orig_topic_compared += 1
            if selected_topic == orig_topic:
                orig_topic_matches += 1
            orig_topic_selected_compared[selected_topic] += 1
            orig_topic_ref_compared[orig_topic] += 1

        if selected_field and orig_field:
            orig_field_compared += 1
            if selected_field == orig_field:
                orig_field_matches += 1
            orig_field_selected_compared[selected_field] += 1
            orig_field_ref_compared[orig_field] += 1

        if (
            selected_topic
            and selected_field
            and orig_topic
            and orig_field
        ):
            orig_joint_compared += 1
            if selected_topic == orig_topic and selected_field == orig_field:
                orig_joint_matches += 1

    def _metric(
        matches: int,
        compared: int,
        left_counter: Optional[Counter[str]] = None,
        right_counter: Optional[Counter[str]] = None,
    ) -> Dict[str, Any]:
        accuracy = (matches / compared) if compared else None
        kappa: Optional[float] = None
        if compared and left_counter is not None and right_counter is not None:
            total = float(compared)
            if total > 0:
                pe = 0.0
                labels = set(left_counter.keys()) | set(right_counter.keys())
                for label in labels:
                    pe += (left_counter.get(label, 0) / total) * (
                        right_counter.get(label, 0) / total
                    )
                po = accuracy if accuracy is not None else 0.0
                denom = 1.0 - pe
                if abs(denom) > 1e-9:
                    kappa = (po - pe) / denom
                    kappa = max(-1.0, min(1.0, kappa))
        return {
            "matches": int(matches),
            "compared": int(compared),
            "mismatches": int(max(0, compared - matches)),
            "accuracy": accuracy,
            "accuracy_percent": round(accuracy * 100.0, 2) if accuracy is not None else None,
            "kappa": round(kappa, 4) if kappa is not None else None,
        }

    def _distribution(counter: Counter[str]) -> Tuple[int, List[Dict[str, Any]]]:
        total = int(sum(counter.values()))
        if total <= 0:
            return 0, []
        top_entries = counter.most_common(8)
        result: List[Dict[str, Any]] = []
        for label, count in top_entries:
            result.append(
                {
                    "label": label,
                    "count": int(count),
                    "percent": round((count / total) * 100.0, 2) if total else 0.0,
                }
            )
        return total, result

    human_topic_total, human_topic_top = _distribution(topic_counter)
    human_field_total, human_field_top = _distribution(field_counter)
    ai_topic_total, ai_topic_top = _distribution(ai_topic_counter)
    ai_field_total, ai_field_top = _distribution(ai_field_counter)

    return {
        "topic": _metric(topic_matches, topic_compared, topic_selected_compared, topic_ai_compared),
        "field": _metric(field_matches, field_compared, field_selected_compared, field_ai_compared),
        "joint": _metric(joint_matches, joint_compared),
        "orig_topic": _metric(
            orig_topic_matches,
            orig_topic_compared,
            orig_topic_selected_compared,
            orig_topic_ref_compared,
        ),
        "orig_field": _metric(
            orig_field_matches,
            orig_field_compared,
            orig_field_selected_compared,
            orig_field_ref_compared,
        ),
        "orig_joint": _metric(
            orig_joint_matches,
            orig_joint_compared,
        ),
        "selected_topic": {
            "total": human_topic_total,
            "unique": len(topic_counter),
            "top": human_topic_top,
        },
        "selected_field": {
            "total": human_field_total,
            "unique": len(field_counter),
            "top": human_field_top,
        },
        "ai_topic": {
            "total": ai_topic_total,
            "unique": len(ai_topic_counter),
            "top": ai_topic_top,
        },
        "ai_field": {
            "total": ai_field_total,
            "unique": len(ai_field_counter),
            "top": ai_field_top,
        },
    }

def _sampling_summary(audit: Dict[str, Any]) -> Dict[str, Any]:
    items = audit.get("items") or []
    total = len(items)
    completed = 0
    partial = 0
    for item in items:
        status = _sampling_item_status(item)
        if status == "completed":
            completed += 1
        elif status == "partial":
            partial += 1
    pending = max(0, total - completed - partial)
    progress = (completed / total) if total else 0.0
    status = "pending"
    if total and completed == total:
        status = "completed"
    elif completed or partial:
        status = "in_progress"
    metrics = _sampling_metrics(audit)
    metrics["progress"] = {
        "completed": completed,
        "total": total,
        "progress": progress,
        "progress_percent": round(progress * 100.0, 2) if total else 0.0,
    }
    return {
        "inspector": audit.get("inspector", ""),
        "inspector_id": audit.get("inspector_id", ""),
        "sample_rate": audit.get("sample_rate", 0.0),
        "sample_rate_percent": round(float(audit.get("sample_rate", 0.0)) * 100.0, 2),
        "sample_size": total,
        "total_rows": int(audit.get("total_rows", 0) or 0),
        "completed": completed,
        "in_progress": partial,
        "pending": pending,
        "progress": progress,
        "progress_percent": round(progress * 100.0, 2),
        "created_at": audit.get("created_at", ""),
        "updated_at": audit.get("updated_at", ""),
        "status": status,
        "seed": audit.get("random_seed"),
        "metrics": metrics,
    }

def _sampling_item_payload(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "item_id": item.get("item_id", ""),
        "row_id": item.get("row_id", ""),
        "row_number": int(item.get("row_number", 0) or 0),
        "sample_order": int(item.get("sample_order", 0) or 0),
        "article_title": item.get("article_title", ""),
        "abstract": item.get("abstract", ""),
        "structured_summary": item.get("structured_summary", ""),
        "ai_topic": item.get("ai_topic", ""),
        "ai_field": item.get("ai_field", ""),
        "orig_topic": item.get("orig_topic", ""),
        "orig_field": item.get("orig_field", ""),
        "selected_topic": item.get("selected_topic", ""),
        "selected_field": item.get("selected_field", ""),
        "checked_at": item.get("checked_at"),
        "status": _sampling_item_status(item),
    }

def _sampling_make_item(row: pd.Series, row_idx: Any, sample_order: int, dataset_position: int) -> Dict[str, Any]:
    return {
        "item_id": str(uuid.uuid4()),
        "row_id": "" if pd.isna(row_idx) else str(row_idx),
        "row_number": int(dataset_position),
        "sample_order": int(sample_order),
        "article_title": safe_get(row, "Article Title"),
        "abstract": safe_get(row, "Abstract"),
        "structured_summary": safe_get(row, "结构化总结"),
        "ai_topic": safe_get(row, ADJUST_TOPIC_COL),
        "ai_field": safe_get(row, ADJUST_FIELD_COL),
        "orig_topic": safe_get(row, "研究主题（议题）分类"),
        "orig_field": safe_get(row, "研究领域分类"),
        "selected_topic": "",
        "selected_field": "",
        "checked_at": None,
    }

def _sampling_normalize_rate(raw: Any) -> float:
    try:
        rate = float(raw)
    except Exception:
        return SAMPLING_DEFAULT_RATE
    if rate <= 0:
        return SAMPLING_DEFAULT_RATE
    if rate > 1:
        rate = rate / 100.0
    if rate <= 0:
        rate = SAMPLING_DEFAULT_RATE
    return min(max(rate, 0.001), 1.0)

def _list_sampling_audits(sid: str) -> List[Dict[str, Any]]:
    directory = os.path.join(_session_dir(sid), SAMPLING_DIR_NAME)
    if not os.path.exists(directory):
        return []
    audits: List[Dict[str, Any]] = []
    for name in os.listdir(directory):
        if not name.endswith(".json"):
            continue
        slug = name[:-5]
        data = _read_sampling_audit(sid, slug)
        if not data:
            continue
        summary = _sampling_summary(data)
        summary["inspector_id"] = data.get("inspector_id", slug)
        summary["_updated_ts"] = float(data.get("updated_ts", 0.0) or 0.0)
        audits.append(summary)
    audits.sort(key=lambda x: x.get("_updated_ts", 0.0), reverse=True)
    for a in audits:
        a.pop("_updated_ts", None)
    return audits

def safe_filename(name:str) -> str:
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = name.strip()
    if not name: name = "export"
    return name

def _normalize_cell_value(value: Any) -> str:
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    lower = text.lower()
    if lower in ("nan", "none", "null", "na"):
        return ""
    return text

def _normalized_series(df: pd.DataFrame, column: str) -> List[str]:
    if column not in df.columns:
        return ["" for _ in range(int(df.shape[0]))]
    series = df[column]
    return [_normalize_cell_value(v) for v in series.tolist()]

def _effective_categories(df: pd.DataFrame, adjust_col: str, fallback_col: str) -> List[str]:
    adjust_vals = _normalized_series(df, adjust_col)
    fallback_vals = _normalized_series(df, fallback_col)
    effective: List[str] = []
    for adj, orig in zip(adjust_vals, fallback_vals):
        val = adj or orig
        if not val:
            val = "未分类"
        effective.append(val)
    return effective

def _extract_year(value: Any) -> Optional[int]:
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if value is None:
        return None
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except Exception:
        m = YEAR_PATTERN.search(text)
        if m:
            try:
                return int(m.group())
            except Exception:
                return None
    return None

def _year_bucket(year: Optional[int], span: int) -> str:
    if not isinstance(span, int) or span < 1:
        span = 1
    if not year:
        return "未知年份"
    start = (int(year) // span) * span
    end = start + span - 1
    return f"{start}-{end}"

# ================== 前端配置 ==================
@app.get("/frontend_config")
def frontend_config():
    return JSONResponse({
        "topic_list": TOPIC_LIST,
        "field_list": FIELD_LIST,
        "adj_topic_key": ADJUST_TOPIC_COL,
        "adj_field_key": ADJUST_FIELD_COL,
        "app_version": APP_VERSION,
        "sampling_default_rate": SAMPLING_DEFAULT_RATE,
        "label_mapping": {
            "topic": dict(TOPIC_LABEL_MAP),
            "field": dict(FIELD_LABEL_MAP),
        },
        "label_mapping_defaults": {
            "topic": dict(DEFAULT_TOPIC_LABEL_MAP),
            "field": dict(DEFAULT_FIELD_LABEL_MAP),
        },
    })

# ================== 上传 / 统计 / 分页 ==================
@app.post("/upload")
def upload(file: UploadFile = File(...)):
    try:
        df = read_table(file)
        sid = str(uuid.uuid4())
        with _get_lock(sid):
            DATASTORE[sid] = df
            _touch(sid)
            save_session_to_disk(sid, origin_filename=file.filename or "")
            _evict_if_needed()
        return JSONResponse({"session_id": sid, "total": int(df.shape[0])})
    except Exception as e:
        return PlainTextResponse(f"读取失败：{e}", status_code=400)

def _timeline(df: pd.DataFrame, bin_years: int) -> Dict[str, Any]:
    try:
        bin_size = int(bin_years)
    except Exception:
        bin_size = 5
    if bin_size < 1:
        bin_size = 1

    if "Publication Year" not in df.columns:
        return {
            "bin_size": bin_size,
            "total": 0,
            "min_year": None,
            "max_year": None,
            "bins": [],
            "stack_topic": {"series": []},
            "stack_field": {"series": []},
        }

    timeline_df = df[["Publication Year", ADJUST_TOPIC_COL, ADJUST_FIELD_COL]].copy()
    timeline_df["Publication Year"] = pd.to_numeric(timeline_df["Publication Year"], errors="coerce")
    timeline_df = timeline_df.dropna(subset=["Publication Year"])
    if timeline_df.empty:
        return {
            "bin_size": bin_size,
            "total": 0,
            "min_year": None,
            "max_year": None,
            "bins": [],
            "stack_topic": {"series": []},
            "stack_field": {"series": []},
        }

    timeline_df = timeline_df[timeline_df["Publication Year"] >= TIMELINE_MIN_YEAR]
    if timeline_df.empty:
        return {
            "bin_size": bin_size,
            "total": 0,
            "min_year": None,
            "max_year": None,
            "bins": [],
            "stack_topic": {"series": []},
            "stack_field": {"series": []},
        }

    timeline_df["Publication Year"] = timeline_df["Publication Year"].astype(int)
    years = timeline_df["Publication Year"]
    min_year = int(years.min())
    max_year = int(years.max())
    # 将区间对齐到 bin_size 的整数倍，便于展示
    start_year = int(math.floor(min_year / bin_size) * bin_size)
    end_year = int(math.ceil((max_year + 1) / bin_size) * bin_size - 1)

    labels: List[str] = []
    total = int(timeline_df.shape[0])

    # 预先计算每行所属的时间区间，便于后续统计
    def _assign_bin(year: int) -> Tuple[int, int]:
        offset = year - start_year
        if offset < 0:
            bucket_index = 0
        else:
            bucket_index = offset // bin_size
        bucket_start = start_year + bucket_index * bin_size
        bucket_end = bucket_start + bin_size - 1
        return bucket_start, bucket_end

    bin_starts: List[int] = []
    bin_ends: List[int] = []
    for start in range(start_year, end_year + 1, bin_size):
        end = start + bin_size - 1
        labels.append(f"{start}-{end}")
        bin_starts.append(start)
        bin_ends.append(end)

    # 将区间标签加入数据帧
    start_series = []
    end_series = []
    label_series = []
    for year in timeline_df["Publication Year"].tolist():
        s, e = _assign_bin(int(year))
        start_series.append(s)
        end_series.append(e)
        label_series.append(f"{s}-{e}")
    timeline_df = timeline_df.assign(
        bin_start=start_series,
        bin_end=end_series,
        bin_label=label_series,
    )

    count_by_bin = timeline_df.groupby("bin_label").size()
    bins: List[Dict[str, Any]] = []
    cumulative = 0
    for idx, label in enumerate(labels):
        start = bin_starts[idx]
        end = bin_ends[idx]
        count = int(count_by_bin.get(label, 0))
        cumulative += count
        percent = (count / total * 100.0) if total else 0.0
        cumulative_percent = (cumulative / total * 100.0) if total else 0.0
        bins.append({
            "label": label,
            "start": int(start),
            "end": int(end),
            "count": count,
            "percent": round(percent, 4),
            "cumulative_count": cumulative,
            "cumulative_percent": round(min(cumulative_percent, 100.0), 4),
        })

    def _build_stack(column: str, empty_label: str, label_mapping: Dict[str, str]) -> Dict[str, Any]:
        cat_series = timeline_df[column].fillna("").astype(str).str.strip()
        cat_series = cat_series.replace("", empty_label)
        stack_df = timeline_df.assign(stack_category=cat_series)
        totals = stack_df.groupby("stack_category").size().sort_values(ascending=False)
        if totals.empty:
            return {"series": []}

        pivot = (
            stack_df.pivot_table(
                index="bin_label",
                columns="stack_category",
                values="Publication Year",
                aggfunc="count",
                fill_value=0,
            )
        )
        pivot = pivot.reindex(labels, fill_value=0)

        series: List[Dict[str, Any]] = []
        for cat, total_count in totals.items():
            if total_count <= 0:
                continue
            counts = pivot[cat].tolist() if cat in pivot.columns else [0] * len(labels)
            series.append({
                "label": str(cat),
                "display_label": _label_display(str(cat), label_mapping),
                "counts": [int(v) for v in counts],
                "total": int(total_count),
            })

        return {
            "series": series,
            "category_total": int(totals.sum()),
            "unique_categories": int(totals.shape[0]),
        }

    return {
        "bin_size": bin_size,
        "total": total,
        "min_year": min_year,
        "max_year": max_year,
        "bins": bins,
        "stack_topic": _build_stack(ADJUST_TOPIC_COL, "未分类", TOPIC_LABEL_MAP),
        "stack_field": _build_stack(ADJUST_FIELD_COL, "未分类", FIELD_LABEL_MAP),
    }


def _render_timeline_line_figure(labels: List[str], counts: List[int]):
    if not labels:
        fig, ax = plt.subplots(figsize=(4.5, 3.5))
        fig.patch.set_facecolor("white")
        ax.axis("off")
        ax.text(
            0.5,
            0.5,
            "No timeline data",
            ha="center",
            va="center",
            fontsize=14,
            color="#1f2937",
            transform=ax.transAxes,
        )
        return fig

    fig_width = min(8.0, max(4.0, len(labels) * 0.6))
    fig_height = min(8.0, max(3.0, fig_width * 0.6))
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    positions = list(range(len(labels)))
    primary_line_color = "#2563eb"
    marker_face_color = "#1d4ed8"

    ax.plot(
        positions,
        counts,
        color=primary_line_color,
        linewidth=2.5,
        marker="o",
        markersize=6,
        markerfacecolor=marker_face_color,
        markeredgewidth=0.9,
        markeredgecolor="#ffffff",
    )

    ax.set_xticks(positions)
    ax.set_xticklabels(labels, rotation=35, ha="right")
    ax.set_xlabel("Year Interval", fontsize=12)
    ax.set_ylabel("Publication Count", fontsize=12)

    max_count = max(counts) if counts else 0
    upper = max_count * 1.1 if max_count > 0 else 1
    ax.set_ylim(0, upper)
    ax.margins(x=0.02)
    ax.set_axisbelow(True)
    ax.grid(axis="y", color="#dbeafe", linewidth=0.8, linestyle="--", alpha=0.7)

    label_offset = max_count * 0.04 if max_count > 0 else 0.5
    for pos, cnt in zip(positions, counts):
        ax.text(
            pos,
            cnt + label_offset,
            f"{cnt:,}",
            ha="center",
            va="bottom",
            fontsize=11,
            color="#1f2937",
            fontweight="semibold",
        )

    for spine_name, spine in ax.spines.items():
        if spine_name in ("left", "bottom"):
            spine.set_visible(True)
            spine.set_color("#1f2937")
            spine.set_linewidth(1.1)
        else:
            spine.set_visible(False)

    ax.tick_params(axis="x", labelsize=11, direction="out", length=4, width=0.8)
    ax.tick_params(axis="y", labelsize=11, direction="out", length=4, width=0.8)

    return fig


@app.get("/category_bar_chart")
def category_bar_chart(
    session_id: str = Query(...),
    dimension: str = Query("topic_adj"),
    format: Optional[str] = Query(None),
    download: bool = Query(False),
):
    dim_key = (dimension or "").strip().lower()
    config = CATEGORY_BAR_CONFIG.get(dim_key)
    if not config:
        return PlainTextResponse(
            "dimension must be one of topic_adj, field_adj, topic_orig, field_orig",
            status_code=400,
        )

    try:
        df = _ensure_session_loaded(session_id)
    except ValueError:
        return PlainTextResponse("无效 session_id", status_code=400)

    column = config["column"]
    if column not in df.columns:
        message = f"数据集中缺少 {column} 列，无法生成图表。"
        if format:
            return PlainTextResponse(message, status_code=404)
        return JSONResponse({"ok": False, "message": message})

    mapping = TOPIC_LABEL_MAP if config["kind"] == "topic" else FIELD_LABEL_MAP
    counts = _counts(df[column], mapping)

    try:
        assets = _generate_category_bar_assets(
            counts,
            dimension=dim_key,
            axis_label=config["axis"],
            title=config["title"],
        )
    except ValueError:
        if format:
            return PlainTextResponse("暂无分类数据", status_code=404)
        return JSONResponse({"ok": False, "message": "暂无分类数据"})

    fmt = (format or "").strip().lower()
    if fmt:
        if fmt not in {"png", "svg"}:
            return PlainTextResponse("format must be png or svg", status_code=400)
        data_url = assets.get(f"{fmt}_data_url")
        if not data_url:
            return PlainTextResponse("暂无图表数据", status_code=404)
        try:
            _, encoded = data_url.split(",", 1)
            binary = base64.b64decode(encoded)
        except Exception:
            return PlainTextResponse("图表生成失败", status_code=500)
        mime = "image/png" if fmt == "png" else "image/svg+xml"
        filename = assets.get(f"{fmt}_filename") or f"{dim_key}_distribution.{fmt}"
        headers = {}
        if download:
            headers["Content-Disposition"] = f"attachment; filename={filename}"
        return Response(content=binary, media_type=mime, headers=headers)

    total_docs = int(sum(int(item.get("count", 0) or 0) for item in counts))
    payload = {
        "ok": True,
        "dimension": dim_key,
        "column": column,
        "title": assets.get("title", config["title"]),
        "total_documents": total_docs,
        "entries": counts,
    }
    payload.update(assets)
    return JSONResponse(payload)


@app.get("/stats")
def stats(session_id: str = Query(...), bin_years: int = Query(5, ge=1, le=50)):
    if session_id not in DATASTORE:
        # 尝试从磁盘加载
        try:
            with _get_lock(session_id):
                DATASTORE[session_id] = load_session_from_disk(session_id)
                _touch(session_id); _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", status_code=400)
    df = DATASTORE[session_id]
    source_col = "Source Title"
    journal_topic = _top_sources_by_category(
        df,
        ADJUST_TOPIC_COL,
        source_col=source_col,
        fallback_category="未分类",
        fallback_source="未提供期刊",
        label_mapping=TOPIC_LABEL_MAP,
    )
    journal_field = _top_sources_by_category(
        df,
        ADJUST_FIELD_COL,
        source_col=source_col,
        fallback_category="未分类",
        fallback_source="未提供期刊",
        label_mapping=FIELD_LABEL_MAP,
    )

    journal_overview = _journal_overview_table(
        df,
        source_col=source_col,
        citation_candidates=["Times Cited, All Databases"],
        quartile_candidates=["JIF Quartile"],
        year_candidates=["Publication Year"],
    )

    return JSONResponse({
        "total": int(df.shape[0]),
        "topic_orig": _counts(df["研究主题（议题）分类"], TOPIC_LABEL_MAP),
        "field_orig": _counts(df["研究领域分类"], FIELD_LABEL_MAP),
        "topic_adj": _counts(df[ADJUST_TOPIC_COL], TOPIC_LABEL_MAP),
        "field_adj": _counts(df[ADJUST_FIELD_COL], FIELD_LABEL_MAP),
        "timeline": _timeline(df, bin_years),
        "journal_source_available": bool(source_col in df.columns),
        "journal_topic_adj": journal_topic,
        "journal_field_adj": journal_field,
        "journal_overview": journal_overview,
    })


@app.get("/timeline_chart")
def timeline_chart(
    session_id: str = Query(...),
    bin_years: int = Query(5, ge=1, le=50),
    format: str = Query("png"),
    download: bool = Query(False),
):
    fmt = format.lower().strip()
    if fmt not in {"png", "svg"}:
        return PlainTextResponse("format must be png or svg", status_code=400)

    if session_id not in DATASTORE:
        try:
            with _get_lock(session_id):
                DATASTORE[session_id] = load_session_from_disk(session_id)
                _touch(session_id)
                _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", status_code=400)

    df = DATASTORE[session_id]
    timeline = _timeline(df, bin_years)
    bins = timeline.get("bins") or []
    labels = [str(item.get("label", "")) for item in bins]
    counts = [int(item.get("count", 0)) for item in bins]

    fig = _render_timeline_line_figure(labels, counts)
    buffer = io.BytesIO()

    if fmt == "png":
        fig.savefig(buffer, format="png", dpi=300, bbox_inches="tight", facecolor="white")
        media_type = "image/png"
        extension = "png"
    else:
        fig.savefig(buffer, format="svg", bbox_inches="tight", facecolor="white")
        media_type = "image/svg+xml"
        extension = "svg"

    plt.close(fig)
    buffer.seek(0)

    filename = f"timeline_{bin_years}yr.{extension}"
    disposition = "attachment" if download else "inline"
    headers = {"Content-Disposition": f"{disposition}; filename={filename}"}

    return Response(buffer.getvalue(), media_type=media_type, headers=headers)


@app.get("/topic_visuals")
def topic_visuals(session_id: str = Query(...)):
    try:
        payload = _get_topic_visual_data(session_id)
    except ValueError:
        return PlainTextResponse("无效 session_id", status_code=400)
    return JSONResponse({"ok": True, **(payload or {})})

@app.get("/word_cloud_assets")
def word_cloud_assets(
    session_id: str = Query(...),
    dimension: str = Query("topic"),
    label: Optional[str] = Query(None),
):
    try:
        payload = _get_topic_visual_data(session_id)
    except ValueError:
        return PlainTextResponse("无效 session_id", status_code=400)

    dim_key = (dimension or "").strip().lower()
    dim_map = {"topic": "topic_adj", "field": "field_adj"}
    if dim_key not in dim_map:
        return PlainTextResponse("dimension 必须为 topic 或 field", status_code=400)

    dim_payload = payload.get(dim_map[dim_key], {}) or {}
    categories = dim_payload.get("categories") or []
    if not categories:
        return JSONResponse({"ok": False, "message": "暂无词云数据"})

    mapping = TOPIC_LABEL_MAP if dim_key == "topic" else FIELD_LABEL_MAP
    target = None
    if label:
        label_str = str(label)
        for cat in categories:
            if str(cat.get("label", "")) == label_str:
                target = cat
                break
    if not target:
        target = categories[0]

    words = target.get("top_words") or []
    try:
        assets = _generate_word_cloud_assets(words)
    except ValueError:
        return JSONResponse({"ok": False, "message": "暂无词频数据"})
    except Exception:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": "词云生成失败"})

    category_label = str(target.get("label") or "未分类")
    category_display = str(target.get("display_label") or _label_display(category_label, mapping))
    slug = _slugify_filename(category_label)
    png_filename = f"{dim_key}_wordcloud_{slug}.png"
    svg_filename = f"{dim_key}_wordcloud_{slug}.svg"

    return JSONResponse(
        {
            "ok": True,
            "category": category_label,
            "category_display": category_display,
            "dimension": dim_key,
            "document_count": int(target.get("count", 0) or 0),
            "total_tokens": int(target.get("total_tokens", 0) or 0),
            "png_filename": png_filename,
            "svg_filename": svg_filename,
            **assets,
        }
    )


@app.get("/keyword_chart_assets")
def keyword_chart_assets(
    session_id: str = Query(...),
    dimension: str = Query("topic"),
    label: Optional[str] = Query(None),
):
    try:
        payload = _get_topic_visual_data(session_id)
    except ValueError:
        return PlainTextResponse("无效 session_id", status_code=400)

    dim_key = (dimension or "").strip().lower()
    dim_map = {"topic": "topic_adj", "field": "field_adj"}
    if dim_key not in dim_map:
        return PlainTextResponse("dimension 必须为 topic 或 field", status_code=400)

    dim_payload = payload.get(dim_map[dim_key], {}) or {}
    categories = dim_payload.get("categories") or []
    if not categories:
        return JSONResponse({"ok": False, "message": "暂无关键词数据"})

    mapping = TOPIC_LABEL_MAP if dim_key == "topic" else FIELD_LABEL_MAP
    target = None
    if label:
        label_str = str(label)
        for cat in categories:
            if str(cat.get("label", "")) == label_str:
                target = cat
                break
    if not target:
        target = categories[0]

    words = target.get("top_words") or []
    try:
        assets = _generate_keyword_bar_assets(words)
    except ValueError:
        return JSONResponse({"ok": False, "message": "暂无关键词数据"})
    except Exception:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": "关键词图生成失败"})

    category_label = str(target.get("label") or "未分类")
    category_display = str(target.get("display_label") or _label_display(category_label, mapping))
    slug = _slugify_filename(category_label)
    png_filename = f"{dim_key}_keywords_{slug}.png"
    svg_filename = f"{dim_key}_keywords_{slug}.svg"

    return JSONResponse(
        {
            "ok": True,
            "category": category_label,
            "category_display": category_display,
            "dimension": dim_key,
            "document_count": int(target.get("count", 0) or 0),
            "total_tokens": int(target.get("total_tokens", 0) or 0),
            "png_filename": png_filename,
            "svg_filename": svg_filename,
            **assets,
        }
    )


@app.get("/data")
def get_data(
    session_id: str = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=5000),
    filter_target: Optional[str] = Query(None),
    filter_value: Optional[str] = Query(None),
    search_scope: Optional[str] = Query(None),
    search_keyword: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),        # "rank"
    sort_order: Optional[str] = Query("asc"),    # "asc"|"desc"
    only_outliers: Optional[int] = Query(0)      # 1|0
):
    if session_id not in DATASTORE:
        try:
            with _get_lock(session_id):
                DATASTORE[session_id] = load_session_from_disk(session_id)
                _touch(session_id); _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", status_code=400)
    df = DATASTORE[session_id]

    # 过滤
    dfv = df
    if filter_target and filter_value:
        if   filter_target=="topic_orig": m = df["研究主题（议题）分类"].astype(str) == filter_value
        elif filter_target=="field_orig": m = df["研究领域分类"].astype(str) == filter_value
        elif filter_target=="topic_adj":  m = df[ADJUST_TOPIC_COL].astype(str)   == filter_value
        elif filter_target=="field_adj":  m = df[ADJUST_FIELD_COL].astype(str)   == filter_value
        else: m = pd.Series([True]*df.shape[0], index=df.index)
        dfv = df[m]

    # 关键词搜索
    keyword = (search_keyword or "").strip()
    if keyword:
        scope = (search_scope or "title").lower()
        if scope == "summary":
            column = "结构化总结"
        else:
            column = "Article Title"
        if column in dfv.columns:
            series = dfv[column].fillna("").astype(str)
            mask = pd.Series(True, index=dfv.index)
            parts = [p for p in re.split(r"\s+", keyword) if p]
            if not parts:
                parts = [keyword]
            for part in parts:
                mask &= series.str.contains(part, case=False, na=False, regex=False)
            dfv = dfv[mask]

    # 只看离群
    if only_outliers and RANK_OUTLIER_COL in dfv.columns:
        dfv = dfv[dfv[RANK_OUTLIER_COL]==True]

    # 排序
    if sort_by=="rank" and RANK_SCORE_COL in dfv.columns:
        dfv = dfv.sort_values(by=[RANK_SCORE_COL], ascending=(str(sort_order).lower()!="desc"), na_position="last")

    total = int(dfv.shape[0])
    total_pages = max(1, math.ceil(total / page_size))
    page = max(1, min(page, total_pages))  # 夹逼

    sub = slice_df(dfv, page, page_size)
    cols = REQUIRED_COLS + [ADJUST_TOPIC_COL, ADJUST_FIELD_COL, RANK_SCORE_COL, RANK_OUTLIER_COL]
    rows = []
    for i, row in sub.iterrows():
        r = {"_index": int(i)}
        for c in cols: r[c] = safe_get(row, c)
        rows.append(r)

    _touch(session_id)
    return JSONResponse({"rows": rows,"page": page,"page_size": page_size,"total": total,"total_pages": total_pages})

# ================== 更新（只改一列）/ 批量 ==================
@app.post("/update")
def update_row(payload: Dict[str, Any]):
    try:
        sid = payload["session_id"]; idx = int(payload["index"])
        if sid not in DATASTORE:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        df = DATASTORE[sid]
        if idx<0 or idx>=df.shape[0]: return PlainTextResponse("行号越界", 400)

        changed = False
        # 结构化总结可独立保存
        if "structured" in payload:
            df.at[idx, "结构化总结"] = payload.get("structured","")
            changed = True

        which = (payload.get("which_adjust","") or "").strip()  # "topic"|"field"
        val   = payload.get("target_value","")

        # 向后兼容（避免同时改两列）
        if not which:
            t_old = payload.get("topic_adjust", None)
            f_old = payload.get("field_adjust", None)
            if (t_old is not None and str(t_old)!="") and (f_old is not None and str(f_old)!=""):
                return PlainTextResponse("一次只能修改一个调整列（topic 或 field）", 400)
            if t_old is not None and str(t_old)!="": which, val = "topic", str(t_old)
            elif f_old is not None and str(f_old)!="": which, val = "field", str(f_old)
            else:
                if changed:
                    DATASTORE[sid] = df
                    save_session_to_disk(sid)
                    _invalidate_topic_viz(sid)
                else:
                    save_session_to_disk(sid)  # 即便只改了结构化总结也落盘
                _touch(sid)
                return JSONResponse({"ok": True})

        if which=="topic":
            if val not in TOPIC_LIST: return PlainTextResponse("目标值不在主题参考列表中", 400)
            df.at[idx, ADJUST_TOPIC_COL] = val
            changed = True
        elif which=="field":
            if val not in FIELD_LIST: return PlainTextResponse("目标值不在领域参考列表中", 400)
            df.at[idx, ADJUST_FIELD_COL] = val
            changed = True
        else:
            return PlainTextResponse("which_adjust 必须是 topic 或 field", 400)

        DATASTORE[sid] = df
        save_session_to_disk(sid)
        if changed:
            _invalidate_topic_viz(sid)
        _touch(sid)
        return JSONResponse({"ok": True})
    except Exception as e:
        return PlainTextResponse(f"更新失败：{e}", 400)

@app.post("/bulk_update_indices")
def bulk_update_indices(payload: Dict[str, Any]):
    sid = payload.get("session_id"); idxs = payload.get("indices",[])
    which = payload.get("which_adjust"); val = payload.get("target_value","")
    if sid not in DATASTORE:
        try:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", 400)
    df = DATASTORE[sid]
    if which not in ("topic","field"): return PlainTextResponse("which_adjust 必须为 topic 或 field", 400)
    if which=="topic" and val not in TOPIC_LIST: return PlainTextResponse("目标值不在主题参考列表中", 400)
    if which=="field" and val not in FIELD_LIST: return PlainTextResponse("目标值不在领域参考列表中", 400)
    n=0
    for i in idxs:
        ii = int(i)
        if 0<=ii<df.shape[0]:
            if which=="topic": df.at[ii, ADJUST_TOPIC_COL] = val
            else: df.at[ii, ADJUST_FIELD_COL] = val
            n+=1
    DATASTORE[sid] = df
    save_session_to_disk(sid)
    if n>0:
        _invalidate_topic_viz(sid)
    _touch(sid)
    return JSONResponse({"ok": True, "updated": n})

@app.post("/bulk_update_filtered")
def bulk_update_filtered(payload: Dict[str, Any]):
    sid = payload.get("session_id"); which = payload.get("which_adjust"); val = payload.get("target_value")
    ft  = payload.get("filter_target"); fv  = payload.get("filter_value")
    if sid not in DATASTORE:
        try:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", 400)
    if which not in ("topic","field"): return PlainTextResponse("which_adjust 应为 topic 或 field", 400)
    if which=="topic" and val not in TOPIC_LIST: return PlainTextResponse("目标值不在主题参考列表中", 400)
    if which=="field" and val not in FIELD_LIST: return PlainTextResponse("目标值不在领域参考列表中", 400)
    df = DATASTORE[sid]
    if   ft=="topic_orig": m = df["研究主题（议题）分类"].astype(str)==fv
    elif ft=="field_orig": m = df["研究领域分类"].astype(str)==fv
    elif ft=="topic_adj":  m = df[ADJUST_TOPIC_COL].astype(str)==fv
    elif ft=="field_adj":  m = df[ADJUST_FIELD_COL].astype(str)==fv
    else: return PlainTextResponse("filter_target 参数不合法", 400)
    idxs = df.index[m].tolist()
    for i in idxs:
        if which=="topic": df.at[i, ADJUST_TOPIC_COL] = val
        else: df.at[i, ADJUST_FIELD_COL] = val
    DATASTORE[sid] = df
    save_session_to_disk(sid)
    if idxs:
        _invalidate_topic_viz(sid)
    _touch(sid)
    return JSONResponse({"ok": True, "updated": len(idxs)})

@app.post("/bulk_copy")
def bulk_copy(session_id: str = Query(...)):
    if session_id not in DATASTORE:
        try:
            with _get_lock(session_id):
                DATASTORE[session_id] = load_session_from_disk(session_id)
                _touch(session_id); _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", 400)
    df = DATASTORE[session_id]
    mt = df[ADJUST_TOPIC_COL].astype(str).str.strip().replace("nan","")== ""
    mf = df[ADJUST_FIELD_COL].astype(str).str.strip().replace("nan","")== ""
    df.loc[mt, ADJUST_TOPIC_COL]  = df.loc[mt, "研究主题（议题）分类"].where(df["研究主题（议题）分类"].isin(TOPIC_LIST), "")
    df.loc[mf, ADJUST_FIELD_COL]  = df.loc[mf, "研究领域分类"].where(df["研究领域分类"].isin(FIELD_LIST), "")
    DATASTORE[session_id] = df
    save_session_to_disk(session_id)
    _invalidate_topic_viz(session_id)
    _touch(session_id)
    return JSONResponse({"ok": True})

# ================== 嵌入 / 排序 ==================
def _ensure_st_model(name: Optional[str]=None):
    global _ST_MODEL, _ST_MODEL_NAME
    if SentenceTransformer is None:
        raise RuntimeError("未安装 sentence-transformers；请改用 OpenAI Embeddings 或安装依赖。")
    if _ST_MODEL is None:
        candidates: List[str] = []
        if name:
            candidates.append(name)
        env_model = os.getenv("SENTENCE_MODEL", "").strip()
        if env_model:
            candidates.append(env_model)
        candidates.extend(_ST_MODEL_FALLBACKS)
        tried = set()
        last_err: Optional[Exception] = None
        for nm in candidates:
            if not nm or nm in tried:
                continue
            tried.add(nm)
            try:
                _ST_MODEL = SentenceTransformer(nm)
                _ST_MODEL_NAME = nm
                break
            except Exception as exc:
                last_err = exc
        if _ST_MODEL is None:
            raise RuntimeError(
                "无法加载可用的 SentenceTransformer 模型，请检查网络或手动设置 SENTENCE_MODEL 环境变量"
            ) from last_err
    return _ST_MODEL, _ST_MODEL_NAME

def embed_local(texts: List[str], model_name: Optional[str], cb=None) -> np.ndarray:
    model,_ = _ensure_st_model(model_name)
    bs = int(os.getenv("EMB_BATCH","64")); total=len(texts); out=[]
    for i in range(0,total,bs):
        v = model.encode(texts[i:i+bs], batch_size=bs, show_progress_bar=False, convert_to_numpy=True, normalize_embeddings=True)
        out.append(v)
        if cb: cb(i+len(v), total)
    return np.vstack(out) if out else np.zeros((0,384),dtype=np.float32)

def embed_openai(texts: List[str], base_url:str, api_key:str, model:str, cb=None) -> np.ndarray:
    url = base_url.rstrip("/") + "/v1/embeddings"
    headers={"Authorization":f"Bearer {api_key}","Content-Type":"application/json"}
    bs = int(os.getenv("OPENAI_EMB_BATCH","100")); total=len(texts); out=[]; done=0
    for i in range(0,total,bs):
        chunk = texts[i:i+bs]
        r = requests.post(url, headers=headers, json={"model":model,"input":chunk}, timeout=120); r.raise_for_status()
        data = r.json()["data"]
        vecs = np.array([np.array(d["embedding"],dtype=np.float32) for d in data])
        vecs = vecs / (np.linalg.norm(vecs,axis=1,keepdims=True)+1e-12)
        out.append(vecs); done += len(chunk)
        if cb: cb(done, total)
    return np.vstack(out) if out else np.zeros((0,1536),dtype=np.float32)

def build_texts(df: pd.DataFrame, fields: Dict[str,bool]) -> List[str]:
    parts = []
    if fields.get("title",True):    parts.append(df["Article Title"].fillna("").astype(str))
    if fields.get("abstract",True): parts.append(df["Abstract"].fillna("").astype(str))
    if fields.get("summary",True):  parts.append(df["结构化总结"].fillna("").astype(str))
    if not parts: parts=[df["Article Title"].fillna("").astype(str)]
    s = parts[0]
    for p in parts[1:]: s = s + " || " + p
    return s.fillna("").astype(str).tolist()

def compute_group_ranking(df: pd.DataFrame, group_by:str, fields:Dict[str,bool], algorithm:str, k:int,
                          src:str, base_url:str, api_key:str, emb_model:str,
                          local_model_name:str, sigma:float, min_cluster_size:int, sid:str) -> Tuple[int,int]:
    gcol = ADJUST_TOPIC_COL if group_by=="topic_adj" else "研究主题（议题）分类"
    mask = df[gcol].astype(str).str.strip()!=""; dfv = df[mask].copy()
    if dfv.empty: return (0,0)

    texts = build_texts(dfv, fields)
    def _emb_cb(done,total): _progress_update(sid,status="embedding",emb_current=done,emb_total=total, message=f"嵌入 {done}/{total}")
    _progress_update(sid,status="embedding",emb_current=0,emb_total=len(texts),groups_done=0,groups_total=0,outliers=0, message="准备嵌入")

    if src=="local":
        emb = embed_local(texts, local_model_name, _emb_cb)
    else:
        if not base_url or not api_key: raise RuntimeError("OpenAI Embeddings 未配置")
        emb = embed_openai(texts, base_url, api_key, emb_model, _emb_cb)

    labels = dfv[gcol].astype(str).tolist()
    idxs   = dfv.index.tolist()
    groups = sorted(set(labels))
    _progress_update(sid,status="grouping",groups_done=0,groups_total=len(groups), message=f"共 {len(groups)} 组")

    total_out = 0
    scores_all = np.zeros(len(idxs), dtype=float)
    outliers_all = np.zeros(len(idxs), dtype=bool)
    cluster_all = np.full(len(idxs), -1, dtype=int)

    for gi, lab in enumerate(groups, start=1):
        sel = [i for i,lb in enumerate(labels) if lb==lab]
        X = emb[sel,:]
        if X.shape[0]==1:
            d = np.array([0.0])
            labs = np.zeros(1, dtype=int)
            outs = np.array([False])
        else:
            if algorithm=="kmeans":
                kk = max(1, min(k, X.shape[0]))
                km = KMeans(n_clusters=kk, n_init=10, random_state=42)
                labs = km.fit_predict(X); centers = km.cluster_centers_
                sims = [float(cosine_similarity(X[i:i+1], centers[labs[i]].reshape(1,-1))[0][0]) for i in range(X.shape[0])]
                d = 1.0 - np.array(sims)
                outs = d > (np.mean(d) + sigma*(np.std(d)+1e-12))
            elif algorithm=="hdbscan":
                if hdbscan is None:
                    raise RuntimeError("未安装 hdbscan，请先 pip install hdbscan")
                mcs = max(2, min_cluster_size)
                clusterer = hdbscan.HDBSCAN(min_cluster_size=mcs, metric="euclidean")
                labs = clusterer.fit_predict(X)
                probs = getattr(clusterer, "probabilities_", np.ones(X.shape[0], dtype=float))
                d = 1.0 - np.array(probs, dtype=float)
                outs = labs == -1
                out_scores = getattr(clusterer, "outlier_scores_", None)
                if out_scores is not None:
                    d = np.maximum(d, np.array(out_scores, dtype=float))
            else:
                cen = X.mean(axis=0, keepdims=True)
                sims= cosine_similarity(X, cen).reshape(-1)
                d = 1.0 - sims
                labs = np.zeros(X.shape[0], dtype=int)
                outs = d > (np.mean(d) + sigma*(np.std(d)+1e-12))

        mu, sd = float(np.mean(d)), float(np.std(d)+1e-12)
        thr = mu + sigma*sd
        if algorithm != "hdbscan":
            outs = d > thr
        total_out += int(np.sum(outs))

        for j, li in enumerate(sel):
            ridx = idxs[li]
            df.at[ridx, RANK_SCORE_COL]   = float(d[j])
            df.at[ridx, RANK_GROUP_COL]   = lab
            df.at[ridx, RANK_OUTLIER_COL] = bool(outs[j])
            if algorithm=="kmeans":
                algo_desc = f"kmeans(k={k})"
            elif algorithm=="hdbscan":
                algo_desc = f"hdbscan(min_cluster_size={mcs})"
            else:
                algo_desc = "centroid"
            df.at[ridx, RANK_ALGO_COL]    = algo_desc
            scores_all[li] = float(d[j])
            outliers_all[li] = bool(outs[j])
            cluster_all[li] = int(labs[j]) if algorithm != "centroid" else 0

        _progress_update(sid,status="grouping",groups_done=gi, message=f"分组 {lab} 完成（{gi}/{len(groups)}）", outliers=total_out)

    try:
        dims = 2
        comps = min(dims, emb.shape[1] if emb.shape[1]>0 else dims, emb.shape[0] if emb.shape[0]>0 else dims)
        coords = np.zeros((emb.shape[0], dims), dtype=float)
        if emb.shape[0] > 0 and comps > 0:
            pca = PCA(n_components=comps)
            trans = pca.fit_transform(emb)
            coords[:, :comps] = trans
        SCATTER_CACHE[sid] = {
            "coords": coords.tolist(),
            "indices": idxs,
            "group_labels": labels,
            "algorithm": algorithm,
            "cluster_labels": cluster_all.tolist(),
            "scores": scores_all.tolist(),
            "outliers": outliers_all.tolist(),
            "group_by": group_by,
            "timestamp": _now_str(),
        }
    except Exception:
        SCATTER_CACHE.pop(sid, None)

    return (int(dfv.shape[0]), int(total_out))

@app.post("/compute_ranking")
def compute_rank(payload: Dict[str,Any]):
    sid = payload.get("session_id")
    try:
        if sid not in DATASTORE:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        df = DATASTORE[sid]
        _progress_init(sid); _progress_update(sid,status="embedding",message="开始计算…")
        SCATTER_CACHE.pop(sid, None)

        total, outs = compute_group_ranking(
            df=df,
            group_by=payload.get("group_by","topic_orig"),
            fields=payload.get("fields",{"title":True,"abstract":True,"summary":True}),
            algorithm=payload.get("algorithm","centroid"),
            k=int(payload.get("k",3)),
            src=payload.get("embedding_source","local"),
            base_url=payload.get("openai_base_url","") or os.getenv("OPENAI_BASE_URL",""),
            api_key =payload.get("openai_api_key","")  or os.getenv("OPENAI_API_KEY",""),
            emb_model=payload.get("openai_emb_model","") or os.getenv("OPENAI_EMBEDDINGS_MODEL","text-embedding-3-large"),
            local_model_name=os.getenv("SENTENCE_MODEL", None),
            sigma=float(payload.get("sigma",2.0)),
            min_cluster_size=int(payload.get("min_cluster_size",5)),
            sid=sid
        )
        DATASTORE[sid] = df
        save_session_to_disk(sid)
        _progress_finish(sid, True, f"完成：样本 {total}，离群 {outs}")
        _touch(sid)
        return JSONResponse({"ok": True, "total": total, "outliers": outs})
    except Exception as e:
        _progress_finish(sid, False, f"失败：{e}")
        return JSONResponse({"detail": f"排序失败：{e}"}, status_code=500)

@app.get("/ranking_scatter")
def ranking_scatter(session_id: str = Query(...)):
    data = SCATTER_CACHE.get(session_id)
    if not data:
        return JSONResponse({"detail": "当前会话暂无排序可视化数据"}, status_code=404)
    return JSONResponse({"ok": True, **data})

# ================== 会话持久化 API ==================
@app.post("/session/save")
def session_save(payload: Dict[str,Any]):
    sid = payload.get("session_id","")
    if not sid: return PlainTextResponse("缺少 session_id", 400)
    if sid not in DATASTORE:
        try:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", 400)
    meta = save_session_to_disk(sid)
    _touch(sid)
    return JSONResponse({"ok": True, "meta": meta})

@app.get("/session/list")
def session_list():
    return JSONResponse({"sessions": list_sessions()})

@app.get("/session/load")
def session_load(session_id: str = Query(...)):
    try:
        with _get_lock(session_id):
            DATASTORE[session_id] = load_session_from_disk(session_id)
            _touch(session_id); _evict_if_needed()
        df = DATASTORE[session_id]
        meta = _read_meta(session_id)
        return JSONResponse({"ok": True, "session_id": session_id, "total": int(df.shape[0]), "meta": meta})
    except Exception as e:
        return PlainTextResponse(f"加载失败：{e}", 400)

@app.get("/session/info")
def session_info(session_id: str = Query(...)):
    meta = _read_meta(session_id)
    if not meta: return PlainTextResponse("未找到会话信息", 404)
    return JSONResponse({"meta": meta})

# ================== 导出 Excel（服务端保存） ==================
def export_excel_to_server(sid:str, file_name:Optional[str]=None) -> str:
    if sid not in DATASTORE:
        with _get_lock(sid):
            DATASTORE[sid] = load_session_from_disk(sid)
            _touch(sid); _evict_if_needed()
    df = DATASTORE[sid]

    # 目标目录
    out_dir = os.path.join(OUTPUT_DIR, sid)
    os.makedirs(out_dir, exist_ok=True)

    # 文件名
    base = safe_filename(file_name or f"export_{_now_str()}")
    if not base.lower().endswith(".xlsx"):
        base += ".xlsx"
    path = os.path.join(out_dir, base)

    # 直接写盘（避免大 BytesIO 占内存）
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Sheet1")
        ws = w.book["Sheet1"]
        from openpyxl.styles import PatternFill
        fill_topic = PatternFill("solid", fgColor="FFF3BF")
        fill_field = PatternFill("solid", fgColor="CDEAFE")
        fill_out   = PatternFill("solid", fgColor="FFDAD6")
        col_idx = {c:i+1 for i,c in enumerate(df.columns.tolist())}
        # 行上色（注意：这里是写盘阶段，不再把整本工作簿放在内存二次复制）
        for r, (_, row) in enumerate(df.iterrows(), start=2):
            if str(row.get(ADJUST_TOPIC_COL,"")) and row.get(ADJUST_TOPIC_COL)!=row.get("研究主题（议题）分类"):
                ws.cell(r, col_idx[ADJUST_TOPIC_COL]).fill = fill_topic
            if str(row.get(ADJUST_FIELD_COL,"")) and row.get(ADJUST_FIELD_COL)!=row.get("研究领域分类"):
                ws.cell(r, col_idx[ADJUST_FIELD_COL]).fill = fill_field
            if bool(row.get(RANK_OUTLIER_COL, False)) and "Article Title" in col_idx:
                ws.cell(r, col_idx["Article Title"]).fill = fill_out

    # 写入 meta
    meta = _read_meta(sid)
    meta["last_export_path"] = path.replace("\\","/")
    meta["last_export_time"] = _now_str()
    _write_meta(sid, meta)

    # 明确触发垃圾回收，降低长时交互内存峰值
    gc.collect()
    return path

@app.post("/export_excel_save")
def export_excel_save(payload: Dict[str,Any]):
    sid = payload.get("session_id","")
    fname = payload.get("file_name","") or None
    if not sid: return PlainTextResponse("缺少 session_id", 400)
    try:
        path = export_excel_to_server(sid, fname)
        # 返回相对路径以及一个可选下载 url
        rel = os.path.relpath(path, start=os.getcwd()).replace("\\","/")
        return JSONResponse({
            "ok": True,
            "saved_path": rel,
            "download_url": f"/file?path={rel}"
        })
    except Exception as e:
        return PlainTextResponse(f"导出失败：{e}", 500)

@app.post("/export_excel_split")
def export_excel_split(payload: Dict[str,Any]):
    sid = payload.get("session_id","")
    targets_raw = payload.get("targets", [])
    if not sid:
        return PlainTextResponse("缺少 session_id", 400)
    if isinstance(targets_raw, str):
        targets = [targets_raw]
    else:
        try:
            targets = list(targets_raw)
        except Exception:
            targets = []
    allowed = {
        "topic": {
            "adjust": ADJUST_TOPIC_COL,
            "fallback": "研究主题（议题）分类",
            "dir": "by_topic",
            "label": "调整后议题"
        },
        "field": {
            "adjust": ADJUST_FIELD_COL,
            "fallback": "研究领域分类",
            "dir": "by_field",
            "label": "调整后领域"
        }
    }
    cleaned: List[str] = []
    for t in targets:
        key = str(t).strip().lower()
        if key in allowed and key not in cleaned:
            cleaned.append(key)
    if not cleaned:
        cleaned = ["topic", "field"]

    try:
        if sid not in DATASTORE:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        df = DATASTORE[sid]

        timestamp = _now_str()
        base_dir = os.path.join(OUTPUT_DIR, sid, f"batch_export_{timestamp}")
        os.makedirs(base_dir, exist_ok=True)

        files: List[Dict[str,Any]] = []
        summary = {key: {"groups": 0, "rows": 0} for key in cleaned}

        for key in cleaned:
            cfg = allowed[key]
            target_dir = os.path.join(base_dir, cfg["dir"])
            os.makedirs(target_dir, exist_ok=True)
            effective = _effective_categories(df, cfg["adjust"], cfg["fallback"])
            group_series = pd.Series(effective, index=df.index, name="_group")
            for label, sub_df in df.groupby(group_series, sort=False):
                group_label = str(label) if label else "未分类"
                if not group_label.strip():
                    group_label = "未分类"
                count = int(sub_df.shape[0])
                summary[key]["groups"] += 1
                summary[key]["rows"] += count
                fname = safe_filename(f"{group_label}({count}篇)")
                path = os.path.join(target_dir, f"{fname}.xlsx")
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    sub_df.to_excel(writer, index=False)
                rel = os.path.relpath(path, start=os.getcwd()).replace("\\","/")
                files.append({
                    "target": key,
                    "label": group_label,
                    "count": count,
                    "path": rel
                })

        meta = _read_meta(sid)
        meta["last_split_export_time"] = timestamp
        meta["last_split_export_dir"] = os.path.relpath(base_dir, start=os.getcwd()).replace("\\","/")
        meta["last_split_targets"] = cleaned
        _write_meta(sid, meta)
        gc.collect()

        return JSONResponse({
            "ok": True,
            "root_dir": meta["last_split_export_dir"],
            "total_files": len(files),
            "targets": cleaned,
            "summary": summary,
            "files": files
        })
    except Exception as e:
        return PlainTextResponse(f"批量导出失败：{e}", 500)

@app.post("/export_topic_year")
def export_topic_year(payload: Dict[str,Any]):
    sid = payload.get("session_id","")
    span_raw = payload.get("year_span") or payload.get("interval") or payload.get("interval_years")
    if not sid:
        return PlainTextResponse("缺少 session_id", 400)
    try:
        span = int(span_raw) if span_raw is not None else 10
    except Exception:
        span = 10
    if span < 1:
        span = 1

    try:
        if sid not in DATASTORE:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        df = DATASTORE[sid]

        effective_topics = _effective_categories(df, ADJUST_TOPIC_COL, "研究主题（议题）分类")
        years_source = _normalized_series(df, "Publication Year")
        years = [_extract_year(v) for v in years_source]

        topic_buckets: Dict[str, Dict[str, List[int]]] = defaultdict(lambda: defaultdict(list))
        for idx, (topic, year_val) in enumerate(zip(effective_topics, years)):
            bucket = _year_bucket(year_val, span)
            topic_buckets[topic][bucket].append(idx)

        timestamp = _now_str()
        base_dir = os.path.join(OUTPUT_DIR, sid, f"topic_year_export_{timestamp}")
        os.makedirs(base_dir, exist_ok=True)

        files: List[Dict[str,Any]] = []
        total_buckets = 0

        for topic, bucket_map in topic_buckets.items():
            topic_label = topic or "未分类"
            topic_dir = os.path.join(base_dir, safe_filename(topic_label) or "topic")
            os.makedirs(topic_dir, exist_ok=True)
            for bucket, indices in bucket_map.items():
                if not indices:
                    continue
                sub_df = df.iloc[indices]
                count = int(sub_df.shape[0])
                total_buckets += 1
                fname = safe_filename(f"{bucket}({count}篇)")
                path = os.path.join(topic_dir, f"{fname}.xlsx")
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    sub_df.to_excel(writer, index=False)
                rel = os.path.relpath(path, start=os.getcwd()).replace("\\","/")
                files.append({
                    "topic": topic_label,
                    "bucket": bucket,
                    "count": count,
                    "path": rel
                })

        meta = _read_meta(sid)
        meta["last_topic_year_export_time"] = timestamp
        meta["last_topic_year_export_dir"] = os.path.relpath(base_dir, start=os.getcwd()).replace("\\","/")
        meta["last_topic_year_span"] = span
        _write_meta(sid, meta)
        gc.collect()

        summary = {
            "topics": len(topic_buckets),
            "span": span,
            "rows": int(df.shape[0]),
            "buckets": total_buckets
        }

        return JSONResponse({
            "ok": True,
            "root_dir": meta["last_topic_year_export_dir"],
            "total_files": len(files),
            "summary": summary,
            "files": files
        })
    except Exception as e:
        return PlainTextResponse(f"议题年份切片导出失败：{e}", 500)

@app.get("/sampling/list")
def sampling_list(session_id: str = Query("")):
    sid = (session_id or "").strip()
    if not sid:
        return JSONResponse({"audits": []})
    try:
        audits = _list_sampling_audits(sid)
        return JSONResponse({"audits": audits})
    except Exception as e:
        return PlainTextResponse(f"拉取抽样列表失败：{e}", status_code=400)

@app.post("/sampling/generate")
def sampling_generate(payload: Dict[str, Any]):
    try:
        sid = (payload.get("session_id") or "").strip()
        inspector = (payload.get("inspector") or "").strip()
        if not sid:
            return PlainTextResponse("缺少 session_id", 400)
        if not inspector:
            return PlainTextResponse("请填写分类核对员姓名", 400)

        force_raw = payload.get("force")
        if isinstance(force_raw, bool):
            force = force_raw
        elif isinstance(force_raw, (int, float)):
            force = bool(force_raw)
        elif isinstance(force_raw, str):
            force = force_raw.strip().lower() in ("1", "true", "yes", "on")
        else:
            force = False

        rate = _sampling_normalize_rate(payload.get("sample_rate", SAMPLING_DEFAULT_RATE))
        try:
            sample_size = int(payload.get("sample_size", 0) or 0)
        except Exception:
            sample_size = 0

        slug = _sampling_slug(inspector)
        existing = _read_sampling_audit(sid, slug)
        if existing and not force:
            summary = _sampling_summary(existing)
            summary["inspector_id"] = existing.get("inspector_id", slug)
            return JSONResponse({"status": "exists", "audit": summary})

        with _get_lock(sid):
            if sid not in DATASTORE:
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid)
                _evict_if_needed()
            df = DATASTORE[sid]
            total_rows = int(df.shape[0])
            if total_rows <= 0:
                return PlainTextResponse("当前会话无数据", 400)
            if sample_size <= 0:
                sample_size = max(1, math.ceil(total_rows * rate))
            sample_size = min(sample_size, total_rows)
            if sample_size <= 0:
                sample_size = 1

            seed = int(time.time())
            rng = random.Random(seed)
            positions = sorted(rng.sample(range(total_rows), sample_size))
            items: List[Dict[str, Any]] = []
            for order, pos in enumerate(positions, start=1):
                row = df.iloc[pos]
                row_label = df.index[pos]
                items.append(_sampling_make_item(row, row_label, order, pos + 1))

            now_ts = time.time()
            now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            audit = {
                "session_id": sid,
                "inspector": inspector,
                "inspector_id": slug,
                "sample_rate": float(rate),
                "sample_size": int(sample_size),
                "total_rows": total_rows,
                "random_seed": seed,
                "created_at": now_text,
                "created_ts": now_ts,
                "updated_at": now_text,
                "updated_ts": now_ts,
                "items": items,
            }
            _write_sampling_audit(sid, slug, audit)
            _touch(sid)

        summary = _sampling_summary(audit)
        summary["inspector_id"] = slug
        return JSONResponse({"status": "created", "audit": summary})
    except Exception as e:
        return PlainTextResponse(f"生成抽样失败：{e}", status_code=400)

@app.get("/sampling/data")
def sampling_data(session_id: str = Query(""), inspector_id: str = Query("")):
    sid = (session_id or "").strip()
    slug = (inspector_id or "").strip()
    if not sid or not slug:
        return PlainTextResponse("缺少必要参数", 400)
    audit = _read_sampling_audit(sid, slug)
    if not audit:
        return PlainTextResponse("未找到抽样任务", 404)
    items_raw = audit.get("items") or []
    items_payload: List[Dict[str, Any]] = []
    for idx, raw in enumerate(items_raw, start=1):
        item_data = dict(raw)
        if not item_data.get("sample_order"):
            item_data["sample_order"] = idx
        if not item_data.get("row_number"):
            try:
                item_data["row_number"] = int(item_data.get("row_id", 0) or 0)
            except Exception:
                item_data["row_number"] = idx
        items_payload.append(_sampling_item_payload(item_data))
    summary = _sampling_summary(audit)
    summary["inspector_id"] = audit.get("inspector_id", slug)
    return JSONResponse({"audit": summary, "items": items_payload})

@app.post("/sampling/save")
def sampling_save(payload: Dict[str, Any]):
    try:
        sid = (payload.get("session_id") or "").strip()
        slug = (payload.get("inspector_id") or "").strip()
        item_id = (payload.get("item_id") or "").strip()
        if not sid or not slug or not item_id:
            return PlainTextResponse("缺少必要参数", 400)

        topic = str(payload.get("topic_choice", "") or "").strip()
        field = str(payload.get("field_choice", "") or "").strip()
        if topic and topic not in TOPIC_LIST:
            return PlainTextResponse("议题选择不在参考列表中", 400)
        if field and field not in FIELD_LIST:
            return PlainTextResponse("领域选择不在参考列表中", 400)

        with _get_lock(sid):
            audit = _read_sampling_audit(sid, slug)
            if not audit:
                return PlainTextResponse("抽样任务不存在", 404)
            items = audit.get("items") or []
            target = None
            for item in items:
                if item.get("item_id") == item_id:
                    target = item
                    break
            if target is None:
                return PlainTextResponse("样本不存在", 404)

            target["selected_topic"] = topic
            target["selected_field"] = field
            if topic or field:
                target["checked_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            else:
                target["checked_at"] = None

            now_ts = time.time()
            audit["items"] = items
            audit["updated_ts"] = now_ts
            audit["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            _write_sampling_audit(sid, slug, audit)
            if sid in DATASTORE:
                _touch(sid)

        summary = _sampling_summary(audit)
        summary["inspector_id"] = audit.get("inspector_id", slug)
        item_payload = _sampling_item_payload(target)
        return JSONResponse({"ok": True, "item": item_payload, "summary": summary})
    except Exception as e:
        return PlainTextResponse(f"保存失败：{e}", status_code=400)

@app.get("/file")
def file_serve(path: str = Query(...)):
    """
    仅允许读取位于 OUTPUT_DIR 或 SESS_DIR 下的文件，防止任意路径下载。
    使用 FileResponse 流式传输，避免内存溢出。
    """
    # 规范化路径
    apath = os.path.abspath(path)
    base_out = os.path.abspath(OUTPUT_DIR)
    base_sess= os.path.abspath(SESS_DIR)
    if not (apath.startswith(base_out) or apath.startswith(base_sess)):
        return PlainTextResponse("路径不允许访问", 403)
    if not os.path.exists(apath):
        return PlainTextResponse("文件不存在", 404)
    filename = os.path.basename(apath)
    return FileResponse(apath, filename=filename, media_type="application/octet-stream")

# ================== LLM 配置 / 健康检查 ==================
@app.get("/default_llm_config")
def default_llm_config():
    return JSONResponse({
        "base_url": os.getenv("OPENAI_BASE_URL",""),
        "api_key":  os.getenv("OPENAI_API_KEY",""),
        "model":    os.getenv("OPENAI_MODEL",""),
        "temp":     0.2
    })

@app.post("/save_env")
def save_env(payload: Dict[str,Any]):
    base=payload.get("base_url","").strip()
    key =payload.get("api_key","").strip()
    model=payload.get("model","").strip()
    try:
        if base:  set_key(".env","OPENAI_BASE_URL",base)
        if key:   set_key(".env","OPENAI_API_KEY",key)
        if model: set_key(".env","OPENAI_MODEL",model)
        if not os.getenv("OPENAI_EMBEDDINGS_MODEL",""):
            set_key(".env","OPENAI_EMBEDDINGS_MODEL","text-embedding-3-large")
        return JSONResponse({"ok": True})
    except Exception as e:
        return PlainTextResponse(f".env 写入失败：{e}", 500)


@app.post("/label_mappings")
def update_label_mappings(payload: Dict[str, Any]):
    try:
        topic_overrides = _sanitize_label_mapping_payload(payload.get("topic"))
        field_overrides = _sanitize_label_mapping_payload(payload.get("field"))
        with LABEL_MAPPING_LOCK:
            _save_label_mapping_file(topic_overrides, field_overrides)
            _update_label_mappings(topic_overrides, field_overrides)
        TOPIC_VIZ_CACHE.clear()
        return JSONResponse(
            {
                "ok": True,
                "label_mapping": {
                    "topic": dict(TOPIC_LABEL_MAP),
                    "field": dict(FIELD_LABEL_MAP),
                },
                "label_mapping_defaults": {
                    "topic": dict(DEFAULT_TOPIC_LABEL_MAP),
                    "field": dict(DEFAULT_FIELD_LABEL_MAP),
                },
            }
        )
    except Exception as e:
        traceback.print_exc()
        return PlainTextResponse(f"保存失败：{e}", status_code=500)

@app.get("/healthz")
def healthz():
    return PlainTextResponse("ok")

# ================== 可选：智能建议占位（防止前端报错） ==================
def _format_topic_prompt(
    row: pd.Series,
    available_topics: List[str],
    retry_hint: Optional[str] = None,
) -> List[Dict[str, str]]:
    """生成用于请求议题分类的提示词。"""
    title = safe_get(row, "Article Title")
    abstract = safe_get(row, "Abstract")
    structured = safe_get(row, "结构化总结")
    content_lines = [
        "请你扮演中文学术分类助手。任务需要使用 GPT 模型依次完成两个步骤：",
        "Step 1：阅读题目与摘要，生成结构化总结。请严格按照以下格式输出四段中文总结，每段不超过60字：",
        "【研究主题】概括研究关注的对象或问题。",
        "【研究方法】说明使用的数据、方法或技术路径。",
        "【核心发现】提炼最重要的结果或结论。",
        "【研究意义】说明成果的价值、应用或启示。",
        "四段内容须按顺序排列，可通过换行分隔。",
        "Step 2：基于完成的结构化总结，从候选列表中严格选择一个最契合的研究主题（议题），不要创造或输出列表外内容。",
        "输出 JSON，字段定义如下：",
        "{",
        "  \"topic_suggestion\": \"候选列表中的研究主题（议题）名称\",",
        "  \"structured_summary\": \"按指定四段格式整理的结构化总结文本\"",
        "}",
        "structured_summary 字段必须返回 Step 1 生成的四段内容。",
        "研究主题（议题）候选列表：" + "、".join(available_topics),
    ]
    if retry_hint:
        content_lines.append(f"上一次回答未通过校验，原因：{retry_hint}。请务必从候选列表中选择正确的主题，并严格返回 JSON。")
    content_lines.extend(
        [
            "请根据以下文献内容完成任务：",
            f"标题：{title or '（无）'}",
            f"摘要：{abstract or '（无）'}",
            f"结构化总结：{structured or '（无）'}",
        ]
    )
    user_prompt = "\n".join(content_lines)
    return [
        {
            "role": "system",
            "content": "You are a helpful assistant that returns strict JSON responses for academic topic classification.",
        },
        {"role": "user", "content": user_prompt},
    ]


def _format_field_prompt(
    row: pd.Series,
    structured_summary: str,
    available_fields: List[str],
    retry_hint: Optional[str] = None,
) -> List[Dict[str, str]]:
    """生成用于请求领域分类的提示词。"""
    title = safe_get(row, "Article Title")
    abstract = safe_get(row, "Abstract")
    structured = structured_summary or safe_get(row, "结构化总结")
    content_lines = [
        "请你扮演中文学术分类助手。根据题目、摘要与结构化总结判断研究领域分类。",
        "结构化总结已按【研究主题】【研究方法】【核心发现】【研究意义】四段整理，可作为核心依据。",
        "必须从候选列表中严格选择一个研究领域，不要创造或输出列表外内容。",
        "输出 JSON，字段定义如下：",
        "{",
        "  \"field_suggestion\": \"候选列表中的研究领域名称\"",
        "}",
        "研究领域候选列表：" + "、".join(available_fields),
    ]
    if retry_hint:
        content_lines.append(f"上一次回答未通过校验，原因：{retry_hint}。请重新从候选列表中选择正确的领域，并严格返回 JSON。")
    content_lines.extend(
        [
            "请根据以下文献内容完成任务：",
            f"标题：{title or '（无）'}",
            f"摘要：{abstract or '（无）'}",
            f"结构化总结：{structured or '（无）'}",
        ]
    )
    user_prompt = "\n".join(content_lines)
    return [
        {
            "role": "system",
            "content": "You are a helpful assistant that returns strict JSON responses for academic field classification.",
        },
        {"role": "user", "content": user_prompt},
    ]


def _prepare_llm_config() -> Tuple[str, str, str]:
    base_url = (os.getenv("OPENAI_BASE_URL") or "").strip()
    api_key = (os.getenv("OPENAI_API_KEY") or "").strip()
    model = (os.getenv("OPENAI_MODEL") or "").strip()
    if not (base_url and api_key and model):
        raise RuntimeError(
            "未配置 LLM 访问参数，请在环境变量或 .env 文件中设置 OPENAI_BASE_URL / OPENAI_API_KEY / OPENAI_MODEL。"
        )
    return base_url, api_key, model


def _invoke_llm(
    messages: List[Dict[str, str]],
    base_url: str,
    api_key: str,
    model: str,
) -> str:
    url = base_url.rstrip("/") + "/chat/completions"
    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0,
        "response_format": {"type": "json_object"},
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    try:
        resp = requests.post(url, headers=headers, data=json.dumps(payload).encode("utf-8"), timeout=60)
    except Exception as e:
        raise RuntimeError(f"请求 LLM 失败：{e}") from e
    if resp.status_code >= 400:
        raise RuntimeError(f"LLM 接口返回错误：{resp.status_code} {resp.text}")
    try:
        data = resp.json()
    except Exception as e:
        raise RuntimeError(f"解析 LLM 返回值失败：{e}") from e
    try:
        content = data["choices"][0]["message"]["content"].strip()
    except Exception as e:
        raise RuntimeError(f"LLM 返回格式不符合预期：{data}") from e
    return content


def _invoke_llm_json(
    messages: List[Dict[str, str]],
    base_url: str,
    api_key: str,
    model: str,
) -> Dict[str, Any]:
    content = _invoke_llm(messages, base_url, api_key, model)
    try:
        return json.loads(content)
    except json.JSONDecodeError as e:
        raise ValueError(f"LLM 返回的内容不是合法 JSON：{content}") from e


def _llm_autosuggest(
    row: pd.Series,
    topic_blacklist: Optional[List[str]] = None,
    field_blacklist: Optional[List[str]] = None,
) -> Dict[str, Any]:
    base_url, api_key, model = _prepare_llm_config()

    topic_blacklist_clean = _sanitize_blacklist(topic_blacklist, TOPIC_LIST)
    field_blacklist_clean = _sanitize_blacklist(field_blacklist, FIELD_LIST)
    topic_blacklist_set = set(topic_blacklist_clean)
    field_blacklist_set = set(field_blacklist_clean)
    available_topics = [t for t in TOPIC_LIST if t not in topic_blacklist_set]
    available_fields = [f for f in FIELD_LIST if f not in field_blacklist_set]
    if not available_topics:
        raise RuntimeError("议题分类黑名单排除了所有候选项，请调整设置。")
    if not available_fields:
        raise RuntimeError("领域分类黑名单排除了所有候选项，请调整设置。")

    topic_retry_hint: Optional[str] = None
    topic_last_error: Optional[str] = None
    topic_suggestion: Optional[str] = None
    structured_summary: str = safe_get(row, "结构化总结")
    for _ in range(2):
        try:
            topic_resp = _invoke_llm_json(
                _format_topic_prompt(row, available_topics, topic_retry_hint),
                base_url,
                api_key,
                model,
            )
        except ValueError as ve:
            topic_last_error = str(ve)
            topic_retry_hint = str(ve)
            continue
        except RuntimeError:
            raise
        topic_candidate = str(topic_resp.get("topic_suggestion", "")).strip()
        summary_candidate = str(topic_resp.get("structured_summary", "")).strip()
        if summary_candidate:
            structured_summary = summary_candidate
        if topic_candidate in available_topics:
            topic_suggestion = topic_candidate
            break
        topic_last_error = f"LLM 议题分类不在候选列表中：{topic_candidate or '（空）'}"
        topic_retry_hint = f"返回的议题“{topic_candidate or '（空）'}”不在候选列表中。"
    if not topic_suggestion:
        raise RuntimeError(topic_last_error or "LLM 未能返回有效的议题分类")

    field_retry_hint: Optional[str] = None
    field_last_error: Optional[str] = None
    field_suggestion: Optional[str] = None
    for _ in range(2):
        try:
            field_resp = _invoke_llm_json(
                _format_field_prompt(
                    row,
                    structured_summary,
                    available_fields,
                    field_retry_hint,
                ),
                base_url,
                api_key,
                model,
            )
        except ValueError as ve:
            field_last_error = str(ve)
            field_retry_hint = str(ve)
            continue
        except RuntimeError:
            raise
        field_candidate = str(field_resp.get("field_suggestion", "")).strip()
        if field_candidate in available_fields:
            field_suggestion = field_candidate
            break
        field_last_error = f"LLM 领域分类不在候选列表中：{field_candidate or '（空）'}"
        field_retry_hint = f"返回的领域“{field_candidate or '（空）'}”不在候选列表中。"
    if not field_suggestion:
        raise RuntimeError(field_last_error or "LLM 未能返回有效的领域分类")

    return {
        "topic_suggestion": topic_suggestion,
        "field_suggestion": field_suggestion,
        "structured_summary": structured_summary or safe_get(row, "结构化总结"),
    }


@app.post("/autosuggest")
def autosuggest(payload: Dict[str,Any]):
    """调用外部 LLM，根据题目/摘要生成分类建议。"""
    sid = payload.get("session_id"); idx = int(payload.get("index", -1))
    if sid not in DATASTORE:
        try:
            with _get_lock(sid):
                DATASTORE[sid] = load_session_from_disk(sid)
                _touch(sid); _evict_if_needed()
        except Exception:
            return PlainTextResponse("无效 session_id", 400)
    df = DATASTORE[sid]
    if idx<0 or idx>=df.shape[0]: return PlainTextResponse("行号越界", 400)
    row = df.iloc[idx]

    topic_blacklist = payload.get("topic_blacklist")
    field_blacklist = payload.get("field_blacklist")

    try:
        result = _llm_autosuggest(row, topic_blacklist=topic_blacklist, field_blacklist=field_blacklist)
    except Exception as e:
        return PlainTextResponse(str(e), 500)

    return JSONResponse(result)

if __name__ == "__main__":
    # Windows 某些环境下 127.0.0.1 权限更稳妥
    uvicorn.run(app, host="127.0.0.1", port=8000)
